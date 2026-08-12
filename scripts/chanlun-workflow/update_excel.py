#!/usr/bin/env python3
"""每日更新Excel: 持仓表+交易记录+心态日志+候选池历史
用法:
  python update_excel.py holdings    # 更新持仓数据
  python update_excel.py pool        # 更新候选池
  python update_excel.py mindset     # 更新心态日志
  python update_excel.py all         # 全部更新
"""
import sys, os, json, shutil
from datetime import datetime, date
from openpyxl import load_workbook
from decimal import Decimal

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
WB = os.path.join(SCRIPT_DIR, 'trade-workbook.xlsx')
DATE = datetime.now()
DATE_STR = DATE.strftime('%Y-%m-%d')

def safe_load_wb():
    from glob import glob
    for p in ['.~lock.*#', '.~lock.*', '~lock.*']:
        for f in glob(os.path.join(SCRIPT_DIR, p)):
            try: os.remove(f)
            except: pass
    return load_workbook(WB)

def safe_save_wb(wb):
    wb.save(WB)
    # 清理锁文件
    from glob import glob
    for p in ['.~lock.*#', '.~lock.*', '~lock.*']:
        for f in glob(os.path.join(SCRIPT_DIR, p)):
            try: os.remove(f)
            except: pass

def get_next_row(ws, col=1):
    for r in range(2, ws.max_row + 2):
        if ws.cell(row=r, column=col).value is None:
            return r
    return ws.max_row + 1

def update_holdings(holdings_data):
    """更新持仓表: 追加当日持仓数据"""
    wb = safe_load_wb()
    ws = wb['持仓表']
    nr = get_next_row(ws)

    for h in holdings_data:
        ws.cell(row=nr, column=1, value=DATE_STR)
        ws.cell(row=nr, column=2, value=h['name'])
        ws.cell(row=nr, column=3, value=h['code'])
        ws.cell(row=nr, column=7, value=h['shares'])
        ws.cell(row=nr, column=8, value=h['entry'])
        ws.cell(row=nr, column=9, value=h['close'])
        ws.cell(row=nr, column=15, value=h['bp_low'])
        ws.cell(row=nr, column=16, value=h['stop'])
        ws.cell(row=nr, column=19, value=h['dl_p'])
        ws.cell(row=nr, column=20, value=True)
        ws.cell(row=nr, column=21, value=h.get('sig_type', '一买'))
        nr += 1

    safe_save_wb(wb)
    print(f"  [持仓表] 已追加 {len(holdings_data)} 条记录")

def update_trade_record(trades):
    """追加交易记录"""
    wb = safe_load_wb()
    ws = wb['交易记录']
    nr = get_next_row(ws)

    for t in trades:
        ws.cell(row=nr, column=1, value=t['date'])
        ws.cell(row=nr, column=2, value=t['name'])
        ws.cell(row=nr, column=3, value=t['code'])
        ws.cell(row=nr, column=4, value=t['direction'])
        ws.cell(row=nr, column=5, value=t.get('sig_type', ''))
        ws.cell(row=nr, column=6, value=t['price'])
        ws.cell(row=nr, column=7, value=t['shares'])
        ws.cell(row=nr, column=8, value=t['amount'])
        nr += 1

    safe_save_wb(wb)
    print(f"  [交易记录] 已追加 {len(trades)} 条记录")

def update_mindset(entry):
    """追加心态日志"""
    wb = safe_load_wb()
    ws = wb['心态日志']
    nr = get_next_row(ws)

    ws.cell(row=nr, column=1, value=DATE_STR)
    ws.cell(row=nr, column=2, value=entry.get('action', ''))
    ws.cell(row=nr, column=3, value=entry.get('chasing', '否'))
    ws.cell(row=nr, column=4, value=entry.get('systematic', '是'))
    ws.cell(row=nr, column=5, value=entry.get('emotion', 3))
    ws.cell(row=nr, column=6, value=entry.get('stop_outside', '否'))
    ws.cell(row=nr, column=7, value=entry.get('trades', 0))
    ws.cell(row=nr, column=8, value=entry.get('compliance', '是'))
    ws.cell(row=nr, column=9, value=entry.get('anxiety', ''))
    ws.cell(row=nr, column=10, value=entry.get('reflection', ''))

    safe_save_wb(wb)
    print(f"  [心态日志] 已追加 1 条记录")

def update_candidate_pool(pool_data):
    """追加候选池历史"""
    wb = safe_load_wb()
    ws = wb['候选池历史']
    nr = get_next_row(ws)

    for p in pool_data:
        ws.cell(row=nr, column=1, value=DATE_STR)
        ws.cell(row=nr, column=2, value=p['code'])
        ws.cell(row=nr, column=3, value=p['name'])
        ws.cell(row=nr, column=4, value=p['price'])
        ws.cell(row=nr, column=5, value=p['ratio'])
        ws.cell(row=nr, column=6, value=p['dl_p'])
        ws.cell(row=nr, column=7, value=p['tier'])
        ws.cell(row=nr, column=8, value=p.get('note', ''))
        nr += 1

    safe_save_wb(wb)
    print(f"  [候选池历史] 已追加 {len(pool_data)} 条记录")

def update_account_summary(total_asset, cash, net_deposit=0, market_state="", signal_grade="",
                          profit_score=None, loss_score=None, trade_count=None,
                          avg_pos=None, avg_days=None, success_rate=None,
                          turnover=None, t0_count=None, t0_rate=None,
                          sector="", style="", beat_index=""):
    """追加账户总表: 只需总资产+可用现金, 公式自动复制"""
    wb = safe_load_wb()
    ws = wb['账户总表']

    # 找最后一个有数据的行 + 1
    nr = 2
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value is not None:
            nr = r + 1

    # 写入日期 + 静态值
    ws.cell(row=nr, column=1, value=DATE_STR)
    ws.cell(row=nr, column=2, value=total_asset)
    ws.cell(row=nr, column=3, value=cash)
    ws.cell(row=nr, column=6, value=net_deposit)

    if market_state:  ws.cell(row=nr, column=10, value=market_state)
    if signal_grade:  ws.cell(row=nr, column=11, value=signal_grade)
    if profit_score is not None:  ws.cell(row=nr, column=12, value=profit_score)
    if loss_score is not None:    ws.cell(row=nr, column=13, value=loss_score)
    if trade_count is not None:   ws.cell(row=nr, column=32, value=trade_count)
    if avg_pos is not None:       ws.cell(row=nr, column=34, value=avg_pos)
    if avg_days is not None:      ws.cell(row=nr, column=35, value=avg_days)
    if success_rate is not None:  ws.cell(row=nr, column=36, value=success_rate)
    if turnover is not None:      ws.cell(row=nr, column=37, value=turnover)
    if t0_count is not None:      ws.cell(row=nr, column=38, value=t0_count)
    if t0_rate is not None:       ws.cell(row=nr, column=39, value=t0_rate)
    if sector:   ws.cell(row=nr, column=40, value=sector)
    if style:    ws.cell(row=nr, column=41, value=style)
    if beat_index: ws.cell(row=nr, column=42, value=beat_index)

    # 从上一行复制公式(行号自动偏移)
    prev = nr - 1
    formula_cols = {4, 5, 7, 8, 9, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25,
                    28, 29, 30, 31, 33, 43, 44, 45, 46, 47, 48, 49, 50, 51}
    for col in formula_cols:
        prev_val = ws.cell(row=prev, column=col).value
        if prev_val and str(prev_val).startswith('='):
            from openpyxl.formula.translate import Translator
            try:
                formula = Translator(prev_val, origin=f"A{prev}").translate_formula(f"A{nr}")
                ws.cell(row=nr, column=col, value=formula)
            except:
                ws.cell(row=nr, column=col, value=prev_val)

    safe_save_wb(wb)
    print(f"  [账户总表] 已追加 {DATE_STR}: 总资产={total_asset}, 现金={cash}")

def clean_duplicates():
    """清理重复行和空行"""
    wb = safe_load_wb()
    total = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 删除空行
        empty = 0
        for r in range(ws.max_row, 1, -1):
            vals = [ws.cell(row=r, column=c).value for c in range(1, 6)]
            if all(v is None for v in vals):
                ws.delete_rows(r, 1)
                empty += 1
        total += empty
    safe_save_wb(wb)
    if total > 0:
        print(f"  [清理] 已删除 {total} 个空行")

if __name__ == '__main__':
    cmd = sys.argv[1] if len(sys.argv) > 1 else 'all'
    if cmd == 'all':
        print("请使用子命令: holdings / pool / mindset / trade / account / clean")
        print("示例: python update_excel.py account 21654.67 9678.67")
    elif cmd == 'account':
        if len(sys.argv) >= 4:
            total = float(sys.argv[2])
            cash = float(sys.argv[3])
            net = float(sys.argv[4]) if len(sys.argv) > 4 else 0
            update_account_summary(total, cash, net)
            print("✅ 账户总表更新完成")
        else:
            print("用法: python update_excel.py account <总资产> <可用现金> [净入金]")
    elif cmd == 'clean':
        clean_duplicates()
        print("✅ 清理完成")
    else:
        print(f"未知命令: {cmd}")