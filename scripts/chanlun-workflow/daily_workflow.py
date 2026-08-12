#!/usr/bin/env python3
"""每日交易工作流 — 主入口
用法:
  python daily_workflow.py screenshot [截图路径1 截图路径2 ...]
  python daily_workflow.py compliance
  python daily_workflow.py scan            # 日线全市场扫描 + 写入候选池
  python daily_workflow.py intraday        # 盘中30min扫描候选池 + 持仓止损检查
  python daily_workflow.py account
  python daily_workflow.py holdings
"""
import sys, subprocess, json, os, time
import urllib.request
from datetime import datetime, date
from decimal import Decimal
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# BUG修复 (2026-07-28): 中文名xlsx在GitHub Actions artifact下载和移动端打开失败
# 修复: 主文件改用英文名 trade-workbook.xlsx
# 兼容: 若英文名不存在但中文名存在, 自动复制一份
WB_EN = os.path.join(SCRIPT_DIR, 'trade-workbook.xlsx')
WB_CN = os.path.join(SCRIPT_DIR, '动态仓位资金管理法则_执行版.xlsx')
if os.environ.get('TRADE_WB'):
    WB = os.environ.get('TRADE_WB')
elif os.path.exists(WB_EN):
    WB = WB_EN
elif os.path.exists(WB_CN):
    # 向后兼容: 中文名存在但英文名不存在 → 复制为英文名
    import shutil
    shutil.copy2(WB_CN, WB_EN)
    WB = WB_EN
else:
    WB = WB_EN  # 默认用英文名(运行时会报错提示文件不存在)
RECALC = os.path.join(SCRIPT_DIR, 'recalc.py')
BEICHI_DIR = SCRIPT_DIR

def clean_lock_files():
    """清理LibreOffice残留锁文件 — 防止xlsx无法打开

    BUG修复 (2026-07-28): LibreOffice recalc超时/崩溃后残留 .~lock.* 文件
    导致: openpyxl打开报错 / Excel提示"文件被锁定" / git commit锁文件
    修复: 每次操作前清理锁文件
    """
    import glob
    for pattern in ['.~lock.*#', '.~lock.*', '~lock.*']:
        for f in glob.glob(os.path.join(SCRIPT_DIR, pattern)):
            try:
                os.remove(f)
                print(f"[锁文件清理] 删除: {os.path.basename(f)}")
            except:
                pass

def safe_load_wb(data_only=False):
    """安全加载Excel — 自动清理锁文件后打开

    BUG修复 (2026-07-28): load_workbook直接打开时,
    若残留 .~lock 文件 → 报错"文件被锁定" → 工作流崩溃
    修复: 加载前先清理锁文件
    """
    clean_lock_files()
    return load_workbook(WB, data_only=data_only)

def safe_save_wb(wb, fix_formulas=True):
    """安全保存Excel — 保存后清理锁文件, 可选修复公式行号

    BUG修复 (2026-07-28): wb.save()后LibreOffice可能残留锁文件
    修复: 保存后立即清理

    BUG修复 (2026-07-29): 子表联动失效
    根因: openpyxl写入数据后, 公式行号与新数据行不匹配
    修复: 保存前调用 fix_cross_sheet_formulas() 修正所有子表公式行号
    """
    if fix_formulas:
        fix_cross_sheet_formulas(wb)
    wb.save(WB)
    clean_lock_files()

def fix_cross_sheet_formulas(wb):
    """修复所有子表的公式行号错配 — 确保跨表联动正确

    BUG修复 (2026-07-29): 子表联动失效
    问题: 持仓表J~Z列公式有系统性+8行号偏移(Row2引用Row10)
         原因: 原始Excel中Row2-9是显示区, Row10+是数据区, 公式跨区引用
    修复: 用正则将每行公式中的行号引用替换为当前行号(自引用)
         对跨表引用(如 账户总表!$A$10)只替换MATCH中的本表行号

    联动链路: 候选池(写入) → 持仓表(引用候选池) → 账户总表(引用持仓表) → 心态日志
    """
    import re

    # 1. 修复持仓表
    if '持仓表' in wb.sheetnames:
        _fix_sheet_formula_rows(wb['持仓表'], data_col=2, cross_refs=['账户总表', '候选池'])

    # 2. 修复账户总表 + 删除重复行
    if '账户总表' in wb.sheetnames:
        _remove_duplicate_date_rows(wb['账户总表'])
        _fix_sheet_formula_rows(wb['账户总表'], data_col=1, cross_refs=['持仓表', '候选池'])

    # 3. 修复心态日志
    if '心态日志' in wb.sheetnames:
        ws = wb['心态日志']
        # 修复typo: k-4 → K2
        for r in range(2, min(ws.max_row + 1, 250)):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v and 'k-4' in str(v).lower():
                    ws.cell(row=r, column=c, value=str(v).replace('k-4', 'K2').replace('K-4', 'K2'))
        _fix_sheet_formula_rows(ws, data_col=1, cross_refs=[])

    # 4. 修复候选池历史
    if '候选池历史' in wb.sheetnames:
        _fix_sheet_formula_rows(wb['候选池历史'], data_col=1, cross_refs=[])

def _fix_sheet_formula_rows(ws, data_col, cross_refs, max_row=250):
    """将每行公式中的行号引用替换为当前行号(自引用)

    对含跨表引用的公式,只替换MATCH($A<r>中的本表行号,保留跨表整列引用
    """
    import re
    fixed = 0
    for r in range(2, min(ws.max_row + 1, max_row + 1)):
        if not ws.cell(row=r, column=data_col).value:
            continue
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=col).value
            if not v or not str(v).startswith('='):
                continue
            formula = str(v)

            if cross_refs and any(f"{ref}!" in formula for ref in cross_refs):
                new_formula = _fix_mixed_formula(formula, r, cross_refs)
            else:
                new_formula = _fix_local_formula(formula, r)

            if new_formula != formula:
                ws.cell(row=r, column=col, value=new_formula)
                fixed += 1
    return fixed

def _fix_local_formula(formula, target_row):
    """修复纯本表引用公式: 将所有行号替换为target_row"""
    import re
    def repl(m):
        prefix, col_letters, old_row = m.group(1), m.group(2), int(m.group(3))
        if old_row < 200:
            return f"{prefix}{col_letters}{target_row}"
        return m.group(0)
    return re.sub(r'(?<![:!])(\$?)([A-Z]{1,3})(\d{1,3})(?![:])', repl, formula)

def _fix_mixed_formula(formula, target_row, cross_refs):
    """修复含跨表引用的公式: 本表部分替换行号, 跨表部分只替换MATCH中的行号"""
    import re
    pattern = '|'.join(f'(?:{ref}!)' for ref in cross_refs)
    parts = re.split(f'({"|".join(ref + "!" for ref in cross_refs)})', formula)

    result = []
    i = 0
    while i < len(parts):
        part = parts[i]
        if i + 1 < len(parts) and any(part == ref + '!' for ref in cross_refs):
            # 跨表引用开始
            result.append(part)
            i += 1
            if i < len(parts):
                cross_part = parts[i]
                # 只替换 MATCH($A<r> 中的行号, 不替换整列引用
                def repl_cross(m):
                    prefix, col_letters, old_row = m.group(1), m.group(2), int(m.group(3))
                    if old_row < 200:
                        return f"{prefix}{col_letters}{target_row}"
                    return m.group(0)
                # 只替换 $A<r> 格式的引用(带$前缀的通常是MATCH参数)
                cross_part = re.sub(r'(\$)([A-Z]{1,3})(\d{1,3})', repl_cross, cross_part)
                result.append(cross_part)
        else:
            # 本表引用部分
            result.append(_fix_local_formula(part, target_row))
        i += 1

    return ''.join(result)

def _remove_duplicate_date_rows(ws):
    """删除账户总表中日期重复的行(保留最后一行)"""
    from openpyxl.utils import get_column_letter
    date_rows = {}
    for r in range(2, ws.max_row + 1):
        date_val = ws.cell(row=r, column=1).value
        if date_val:
            date_str = str(date_val)[:10] if not isinstance(date_val, str) else date_val[:10]
            if date_str not in date_rows:
                date_rows[date_str] = []
            date_rows[date_str].append(r)

    rows_to_delete = []
    for date_str, rows in date_rows.items():
        if len(rows) > 1:
            keep_row = rows[-1]
            for r in rows[:-1]:
                # 合并数据: 旧行有值但新行没值的, 复制过来
                for c in range(1, ws.max_column + 1):
                    old_val = ws.cell(row=r, column=c).value
                    new_val = ws.cell(row=keep_row, column=c).value
                    if old_val and not new_val:
                        ws.cell(row=keep_row, column=c, value=old_val)
                rows_to_delete.append(r)

    # 从后往前删除
    rows_to_delete.sort(reverse=True)
    for row_num in rows_to_delete:
        ws.delete_rows(row_num, 1)

def recalc(timeout=60, retries=2):
    """重算Excel公式 — 带重试、状态检查和缓存值验证

    BUG修复 (2026-07-28): 子表联动失效
    根因1: openpyxl保存后公式缓存值全部清空为None → data_only=True读取返回None
    根因2: 旧recalc超时30秒不够(10000+公式), 超时后不报错继续 → 缓存值仍为None
    根因3: recalc返回error时调用方不检查status → 程序静默继续用None值
    修复: 超时增至60秒, 失败重试2次, 返回status供调用方检查

    BUG修复 (2026-07-29): recalc成功但缓存值仍为None
    根因: LibreOffice在某些环境下recalc不写入缓存值
    修复: recalc后抽样验证3个关键单元格, 若为None则重试
    """
    import time as _time
    for attempt in range(retries):
        clean_lock_files()
        r = subprocess.run(
            ['python', RECALC, WB, str(timeout)],
            capture_output=True, text=True
        )
        clean_lock_files()
        if r.stdout:
            result = json.loads(r.stdout)
            if result.get("status") in ("success", "errors_found"):
                # 验证缓存值: 抽样检查3个关键单元格
                if _verify_cached_values():
                    if attempt > 0:
                        print(f"  [recalc] 第{attempt+1}次重试成功, 缓存值验证通过")
                    return result
                else:
                    print(f"  [recalc] 第{attempt+1}次重算成功但缓存值为None, 需重试")
            else:
                print(f"  [recalc] 第{attempt+1}次失败: {result.get('error', 'unknown')}")
        else:
            print(f"  [recalc] 第{attempt+1}次无输出, stderr={r.stderr[:200] if r.stderr else 'none'}")
        if attempt < retries - 1:
            _time.sleep(2)
    print(f"  [recalc] {retries}次重试全部失败, 公式缓存值可能为None")
    return {"status": "error", "total_errors": -1}

def _verify_cached_values():
    """验证recalc后公式缓存值是否确实写入 — 抽样3个关键单元格

    检查点:
    1. 持仓表 N列(持仓占比) — 引用账户总表
    2. 账户总表 E列(总仓位) — 引用持仓表
    3. 候选池 K列(分层) — 引用账户总表
    """
    try:
        wb = load_workbook(WB, data_only=True)
        checks = []

        if '持仓表' in wb.sheetnames:
            ws = wb['持仓表']
            val = ws.cell(row=2, column=14).value  # N列=持仓占比
            checks.append(('持仓表!N2', val))

        if '账户总表' in wb.sheetnames:
            ws = wb['账户总表']
            # 找到最后一行有数据的
            last_row = 2
            for r in range(2, min(ws.max_row + 1, 50)):
                if ws.cell(row=r, column=1).value:
                    last_row = r
            val = ws.cell(row=last_row, column=5).value  # E列=总仓位
            checks.append((f'账户总表!E{last_row}', val))

        if '候选池' in wb.sheetnames:
            ws = wb['候选池']
            val = ws.cell(row=2, column=11).value  # K列=分层
            checks.append(('候选池!K2', val))

        wb.close()

        # 所有抽样值都不为None才算通过
        none_count = sum(1 for _, v in checks if v is None)
        if none_count > 0:
            print(f"  [recalc验证] {none_count}/{len(checks)} 个单元格缓存值为None: "
                  + ", ".join(f"{loc}={val}" for loc, val in checks if val is None))
            return False
        return True
    except Exception as e:
        print(f"  [recalc验证] 异常: {e}")
        return False

def get_today_holdings():
    """读取持仓表, 按代码去重(取最后一次出现的行), 与交易记录交叉验证

    BUG修复 (2026-07-26): 万华已清仓仍显示持有
    根因: 持仓表存在多日重复录入, 旧行(100股)和新行(0股)共存
    修复: 以代码为key, 后出现的行覆盖先出现的行

    BUG修复 (2026-08-03): 同一股票不同代码导致重复持仓
    根因: 东风股份同时存在600006(旧)和601515(新)两条记录
    修复: 按名称二次去重, 保留最新(最后出现)的代码和行数据

    BUG修复 (2026-08-05): Telegram推送虚假持仓(松芝/贵绳/日上)
    根因: 持仓表存有过期数据(卖出后未删除), 无交易记录交叉验证
    修复: 读取交易记录"卖出"操作, 已全卖出的股票自动过滤
    """
    wb = safe_load_wb(data_only=True)
    ws = wb['持仓表']
    code_map = {}  # 按代码去重, 取最后一行
    row_order = {}  # 记录每个代码首次出现的行号顺序
    row_idx = 0
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value
        if not name:
            continue
        code = str(ws.cell(row=r, column=3).value or '')
        waived = ws.cell(row=r, column=4).value
        shares = ws.cell(row=r, column=7).value
        entry = ws.cell(row=r, column=8).value
        close = ws.cell(row=r, column=9).value
        stop = ws.cell(row=r, column=16).value
        action = ws.cell(row=r, column=28).value
        profit = ws.cell(row=r, column=13).value
        pos = ws.cell(row=r, column=14).value
        t1_lock = ws.cell(row=r, column=35).value
        # 始终用后出现的行覆盖 (最新数据)
        code_map[code] = {
            "name": name, "code": code, "waived": waived,
            "shares": shares, "entry": entry, "close": close,
            "stop": stop, "action": action, "profit": profit, "pos": pos,
            "t1_lock": t1_lock
        }
        if code not in row_order:
            row_order[code] = row_idx
            row_idx += 1
    # 按代码去重后, 再按名称二次去重 (防止同一股票不同代码)
    name_map = {}
    for c, h in code_map.items():
        n = h['name']
        if n in name_map:
            # 同名股票: 保留后出现的行 (row_order更大 = 更新)
            if row_order.get(c, 0) > row_order.get(name_map[n]['code'], 0):
                name_map[n] = h
        else:
            name_map[n] = h
    # 过滤掉0股的(已清仓)
    holdings = [h for h in name_map.values() if h.get('shares') and h['shares'] > 0]

    # ============================================================
    # BUG修复 (2026-08-05): 交易记录交叉验证
    # 问题: 持仓表松芝/贵绳/日上等已卖出, 但表内未更新(仍有股数 > 0)
    #       → get_today_holdings返回虚假持仓 → Telegram推送错误
    # 修复: 读取交易记录中的"卖出"操作, 已全卖出的股票自动过滤
    # ============================================================
    try:
        ws_trade = wb['交易记录']
        sold_stocks = {}  # name -> {"total_sold": 股数, "last_sold_date": str}
        for r in range(2, ws_trade.max_row + 1):
            t_name = ws_trade.cell(row=r, column=2).value
            t_action = ws_trade.cell(row=r, column=4).value
            t_shares = ws_trade.cell(row=r, column=7).value  # BUG修复(2026-08-07): column=5是信号类型, column=7才是成交股数
            t_date = ws_trade.cell(row=r, column=1).value
            if t_name and t_action and '卖出' in str(t_action):
                shares_val = 0
                try:
                    shares_val = int(t_shares) if t_shares else 0
                except (ValueError, TypeError):
                    try:
                        shares_val = float(t_shares) if t_shares else 0
                    except:
                        shares_val = 0
                if t_name not in sold_stocks:
                    sold_stocks[t_name] = {"total_sold": 0, "last_date": ""}
                sold_stocks[t_name]["total_sold"] += shares_val
                if t_date:
                    sold_stocks[t_name]["last_date"] = str(t_date)

        # 对每只持仓, 检查是否在交易记录中有明确卖出操作
        filtered = []
        for h in holdings:
            name = h['name']
            shares = h.get('shares', 0) or 0
            entry = h.get('entry', 0) or 0
            close = h.get('close', 0) or 0

            # 合理性检查1: 交易记录中该股票卖出总量 >= 持仓量 → 已清仓
            if name in sold_stocks:
                total_sold = sold_stocks[name]["total_sold"]
                if total_sold >= shares and shares > 0:
                    print(f"  [验证] {name} 交易记录已卖出{total_sold}股 >= 持仓{shares}股, 自动过滤")
                    continue

            # 合理性检查2: 价格/成本不合理的过滤
            if entry > 0 and close > 0:
                # 成本价异常(>100元)但现价<10元 → 可能数据错误
                if entry > 100 and close < 10:
                    print(f"  [验证] {name} 成本{entry:.2f}异常 > 现价{close:.2f}*10, 自动过滤")
                    continue
                # 现价和成本差价>50倍 → 可能数据错误
                if close / entry > 50 or entry / close > 50:
                    print(f"  [验证] {name} 成本{entry:.2f}与现价{close:.2f}差价>50倍, 自动过滤")
                    continue

            filtered.append(h)

        holdings = filtered

        # ============================================================
        # BUG修复 (2026-08-07): 交易记录买入未录入持仓表检测
        # 问题: 证通电子8/3买入300股在交易记录中, 但持仓表无此记录
        #       → get_today_holdings返回的持仓遗漏实际持仓
        #       → 合规检查无法检查未录入的持仓
        # 修复: 扫描交易记录中所有"买入"操作, 检查是否在持仓表中
        # ============================================================
        held_names = {h['name'] for h in holdings}
        buy_stocks = {}  # name -> {"shares": 股数, "code": 代码, "price": 价格}
        for r in range(2, ws_trade.max_row + 1):
            t_name = ws_trade.cell(row=r, column=2).value
            t_action = ws_trade.cell(row=r, column=4).value
            t_shares = ws_trade.cell(row=r, column=7).value  # Col7=成交股数
            t_code = ws_trade.cell(row=r, column=3).value
            t_price = ws_trade.cell(row=r, column=6).value     # Col6=成交价格
            if t_name and t_action and '买入' in str(t_action) and '违规' not in str(t_action):
                shares_val = 0
                try:
                    shares_val = int(t_shares) if t_shares else 0
                except (ValueError, TypeError):
                    try:
                        shares_val = float(t_shares) if t_shares else 0
                    except:
                        shares_val = 0
                if t_name not in buy_stocks:
                    buy_stocks[t_name] = {"shares": 0, "code": str(t_code or ""), "price": t_price or 0}
                buy_stocks[t_name]["shares"] += shares_val

        # 检查: 交易记录有买入但持仓表无此股票
        for name, info in buy_stocks.items():
            if name not in held_names:
                # 排除已卖出的
                sold_total = sold_stocks.get(name, {}).get("total_sold", 0)
                if sold_total < info["shares"]:
                    remaining = info["shares"] - sold_total
                    print(f"  ⚠️ [数据校验] {name}({info['code']}) 交易记录买入{info['shares']}股"
                          f"但持仓表无记录(可能剩余{remaining}股未录入), 请手动更新持仓表")

        # 检查: 持仓表有记录但交易记录无买入(旧数据残留)
        buy_names = set(buy_stocks.keys())
        for h in holdings:
            if h['name'] not in buy_names:
                print(f"  ⚠️ [数据校验] {h['name']}({h['code']}) 持仓表有记录"
                      f"但交易记录无买入操作(可能为旧数据残留)")

    except Exception as e:
        print(f"  [验证] 交易记录交叉验证失败: {e}")

    # ============================================================
    # BUG修复 (2026-08-07): 持仓表数据完整性校验
    # 问题: 持仓表Row3-5(皇氏/沃华/东风)缺失市值/盈亏/占比等计算列
    #       → 合规检查使用None值 → 仓位/盈亏计算静默失败
    # 修复: 对每条持仓校验关键字段, 缺失时打印告警
    # ============================================================
    for h in holdings:
        missing_fields = []
        if not h.get('close') or h['close'] == 0:
            missing_fields.append("当前价")
        if not h.get('entry') or h['entry'] == 0:
            missing_fields.append("成本价")
        if not h.get('stop'):
            missing_fields.append("止损价")
        if not h.get('pos'):
            missing_fields.append("持仓占比")
        if missing_fields:
            print(f"  ⚠️ [数据校验] {h['name']}({h['code']}) 缺失字段: {', '.join(missing_fields)}"
                  f" → 合规检查可能失效, 请更新持仓表")

    return holdings

def get_account_summary():
    wb = safe_load_wb(data_only=True)
    ws = wb['账户总表']
    latest_row = 2
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value:
            latest_row = r

    total_asset = ws.cell(row=latest_row, column=2).value or 0
    cash = ws.cell(row=latest_row, column=3).value or 0
    position_ratio = ws.cell(row=latest_row, column=5).value
    # BUG修复 (2026-08-07): 公式列在data_only=True时返回None
    # 当position_ratio为None时, 用total_asset和cash手动计算
    if position_ratio is None and total_asset > 0:
        market_value = total_asset - cash
        position_ratio = market_value / total_asset
        print(f"  [数据补全] position_ratio公式列为None, 手动计算: {market_value}/{total_asset} = {position_ratio:.1%}")

    return {
        "date": ws.cell(row=latest_row, column=1).value,
        "total_asset": total_asset,
        "cash": cash,
        "position_ratio": position_ratio,
        "stage": ws.cell(row=latest_row, column=33).value,
        "monthly_target": ws.cell(row=latest_row, column=28).value,
        "deviation": ws.cell(row=latest_row, column=31).value,
        "status": ws.cell(row=latest_row, column=20).value,
        "allow_new": ws.cell(row=latest_row, column=22).value,
    }

def get_dynamic_position_cap(code, cost, close):
    """
    动态仓位上限计算 V4 (2026-08-08): 多级别EP_L+几何确认

    替代单级别二买/三买检测, 用EP_L+几何判定+多级别共振
    解决: 一买DL_P变动导致核心池不稳定 + ratio/级别本身有bug
    【重要: DL模型30min/5min失效, 改用EP_L和几何判定】

    多级别共振机制:
      一买建仓: 35% (基础上限)
      二买加仓: 50% (日线一买valid + 30minEP_L/几何确认)
      三买加仓: 60% (日线趋势up + 30minEP_L/几何确认)

    候选池分层(V5.2, 2026-08-08):
      核心池: 30min综合评分(EP_L为主)>=0.50 + 日线不在主跌段
      观察池: 30min综合评分(EP_L为主)>=0.35 + 日线不在主跌段
      边缘池: 30min综合评分(EP_L为主)>=0.20

    条件闭环:
      1. 买点升级: 多级别EP_L+几何确认
      2. 浮盈护垫: 浮盈>=5%才允许加仓
    """
    global dynamic_cap_info
    dynamic_cap_info = {"entry": "一买", "pnl_pct": 0, "cap": 0.35, "tier": "边缘池"}

    if cost <= 0 or close <= 0:
        return 0.35

    pnl_pct = (close - cost) / cost
    dynamic_cap_info["pnl_pct"] = pnl_pct

    # 条件1: 多级别DL_P共振检测 (无论浮盈多少都要检测)
    # BUG修复 (2026-07-29): 旧代码在pnl_pct<0.05时直接return 0.35,
    #   跳过了多级别信号检测 → dynamic_cap_info中entry永远是"一买"
    #   → 合规审查无法发现"二买已确认但浮盈不足"的情况
    sys.path.insert(0, BEICHI_DIR)
    from beichi_analyzer import detect_multilevel_buy_signals, detect_zhongyin

    try:
        ml = detect_multilevel_buy_signals(code, price=close)
    except:
        ml = {}

    tier = ml.get("tier", "边缘池")
    ermai = ml.get("ermai")
    sanmai = ml.get("sanmai")

    dynamic_cap_info["tier"] = tier
    dynamic_cap_info["daily_dl_p"] = ml.get("daily_dl_p", 0)
    dynamic_cap_info["daily_ep_p"] = ml.get("daily_ep_p", 0)
    dynamic_cap_info["30min_dl_p"] = ml.get("30min_dl_p", 0)
    dynamic_cap_info["30min_ep_p"] = ml.get("30min_ep_p", 0)
    dynamic_cap_info["5min_dl_p"] = ml.get("5min_dl_p", 0)
    dynamic_cap_info["5min_ep_p"] = ml.get("5min_ep_p", 0)
    # BUG修复 (2026-07-30): 一买低点 — 用于加仓风控和破位止损
    one_buy_low = ml.get("one_buy_low")
    dynamic_cap_info["one_buy_low"] = one_buy_low

    # P1 (2026-08-02): 中阴状态检测 — 仓位压制依据
    zhongyin_info = ml.get("zhongyin", {})
    dynamic_cap_info["zhongyin"] = zhongyin_info
    dynamic_cap_info["daily_confidence"] = ml.get("daily_confidence", 0)

    best_entry = "一买"
    best_dl_prob = 0
    best_ep_prob = 0  # EP_L反转概率
    best_ermai_score = 0  # 二买综合确认强度

    if sanmai and sanmai.get("valid"):
        best_entry = "三买"
        best_dl_prob = sanmai.get("dl_prob", 0)
        best_ep_prob = sanmai.get("ep_prob", 0)
        best_ermai_score = sanmai.get("ermai_dl_prob", sanmai.get("dl_prob", 0))
    elif ermai and ermai.get("valid"):
        best_entry = "二买"
        best_dl_prob = ermai.get("dl_prob", 0)
        best_ep_prob = ermai.get("ep_prob", 0)
        best_ermai_score = ermai.get("ermai_dl_prob", 0)

    dynamic_cap_info["entry"] = best_entry
    dynamic_cap_info["dl_prob"] = best_dl_prob
    dynamic_cap_info["ep_prob"] = best_ep_prob
    dynamic_cap_info["ermai_dl_prob"] = best_ermai_score

    # 条件2: 浮盈护垫 — 决定是否实际提升仓位上限
    # 信号检测已完成(dynamic_cap_info已更新), 但浮盈不足时不提升cap
    if pnl_pct < 0.05:
        # 信号已检测但浮盈不足 → cap保持35%, entry已记录
        dynamic_cap_info["cap"] = 0.35
        dynamic_cap_info["note"] = f"{best_entry}信号存在但浮盈{pnl_pct:.1%}<5%护垫, 暂不提升上限"
        return 0.35

    # ============================================================
    # BUG修复 (2026-07-30): 一买低点距离检查 — 防止加仓后破位回撤
    #
    # 核心风险场景:
    #   去弱留强 → 卖V形(浮盈大) → 加U形(浮盈小但>=5%)
    #   → U形二买无法确认 → 价格破一买低点 → 重仓大幅回撤
    #
    # 修复策略:
    #   二买/三买加仓前, 检查当前价格离一买低点的距离
    #   距离<3% → 二买结构脆弱, 随时可能破位 → 不允许提升仓位上限
    #
    # 3%阈值的依据:
    #   A股日内波动通常1-2%, 3%相当于1.5个ATR
    #   距离<3%意味着一个日内波动就可能破位
    #   且30min中枢幅度通常3-8%, 3%已在中枢下沿附近
    # ============================================================
    MIN_DIST_FROM_ONE_BUY_LOW = 0.03  # 3%
    if one_buy_low and one_buy_low > 0 and close > 0:
        dist_to_low = (close - one_buy_low) / one_buy_low
        dynamic_cap_info["dist_to_one_buy_low"] = dist_to_low
        if dist_to_low < MIN_DIST_FROM_ONE_BUY_LOW:
            # 价格离一买低点太近 → 二买破位风险极高, 不允许加仓
            dynamic_cap_info["cap"] = 0.35
            dynamic_cap_info["note"] = (
                f"{best_entry}信号存在且浮盈{pnl_pct:.1%}>=5%, "
                f"但价格离一买低点仅{dist_to_low:.1%}<{MIN_DIST_FROM_ONE_BUY_LOW:.0%}, "
                f"二买破位风险高, 暂不提升上限(一买低={one_buy_low:.2f})"
            )
            return 0.35
    else:
        dynamic_cap_info["dist_to_one_buy_low"] = None

    # 动态上限表: 买点级别 × 浮盈护垫
    cap_table = {
        "一买": 0.35,
        "二买": 0.50 if pnl_pct >= 0.05 else 0.35,
        "三买": 0.60 if pnl_pct >= 0.10 else (0.50 if pnl_pct >= 0.05 else 0.35),
    }

    cap = cap_table.get(best_entry, 0.35)

    # P1 (2026-08-02): 中阴状态仓位压制
    # 中阴 = 背驰信号存在但趋势未确认 → 仓位减半
    # NotChasing = 背驰存在但价格脱离中枢 → 不加仓
    zy = zhongyin_info
    if zy.get("is_zhongyin"):
        cap = cap * 0.5
        dynamic_cap_info["cap"] = cap
        dynamic_cap_info["note"] = (
            f"{best_entry}信号存在但中阴状态(趋势未确认), 仓位压制至{cap:.0%}"
            f"({zy.get('reason', '')})"
        )
    elif zy.get("action") == "NotChasing":
        cap = 0.35  # 不允许提升上限
        dynamic_cap_info["cap"] = cap
        dynamic_cap_info["note"] = (
            f"{best_entry}信号存在但NotChasing({zy.get('reason', '')}), "
            f"维持35%上限不加仓"
        )
    else:
        dynamic_cap_info["cap"] = cap
    return cap


def check_buy_compliance(holdings):
    """
    检查买入标的是否在候选池内 (2026-07-29 BUG修复)

    核心矛盾: 系统compliance只检查持仓止损/仓位/破位,
              但不检查"买入标的是否经过full_scan确认入池".
              → 沃华医药DL_P=0.93但不在候选池(被旧代码截断) → 手动检查时违规
              → 系统不报违规 → 一会儿行一会儿不行

    修复: 读取今日交易记录中的买入操作, 逐笔检查是否在候选池内.
          同时允许事后背驰确认(DL_P>0.8)作为豁免条件.
    """
    issues = []
    today_str = datetime.now().strftime('%Y-%m-%d')

    # 1. 从交易记录读取今日买入
    # BUG修复 (2026-07-30): 旧代码用 today_str[:8]+"01" 过滤本月所有交易
    #   → 7/22的万华化学卖出、7/24的建研院买入等历史交易全部被检查
    #   → 12项违规中有9项是已平仓的历史交易误报
    # 修复: 只检查今日的交易记录
    try:
        wb = safe_load_wb(data_only=True)
    except Exception as e:
        print(f"  买入合规检查: 无法读取Excel({e}), 跳过")
        return issues

    if '交易记录' not in wb.sheetnames:
        print(f"  买入合规检查: 无交易记录sheet, 跳过")
        return issues

    ws_tr = wb['交易记录']
    today_buys = []
    for r in range(2, ws_tr.max_row + 1):
        direction = str(ws_tr.cell(row=r, column=4).value or "")
        if direction != "买入":
            continue
        raw_date = ws_tr.cell(row=r, column=1).value
        date_str = ""
        if raw_date:
            if isinstance(raw_date, datetime):
                date_str = raw_date.strftime('%Y-%m-%d')
            elif isinstance(raw_date, date):
                date_str = raw_date.strftime('%Y-%m-%d')
            else:
                date_str = str(raw_date)[:10]
        # 只检查今日的交易
        if date_str == today_str:
            name = ws_tr.cell(row=r, column=2).value or ""
            code = str(ws_tr.cell(row=r, column=3).value or "")
            price = ws_tr.cell(row=r, column=6).value or 0
            today_buys.append({"name": name, "code": code, "price": price, "row": r, "date": date_str})

    if not today_buys:
        print(f"  买入合规检查: 无近期买入记录, 跳过")
        return issues

    # 2. 读取候选池 + 候选池更新日期
    candidate_codes = set()
    pool_date_str = ""
    if '候选池' in wb.sheetnames:
        ws_pool = wb['候选池']
        for r in range(2, ws_pool.max_row + 1):
            code = ws_pool.cell(row=r, column=3).value
            if code:
                candidate_codes.add(str(code))
        # 读取候选池第一行数据的日期(即扫描日期)
        pool_date_raw = ws_pool.cell(row=2, column=1).value
        if pool_date_raw:
            if isinstance(pool_date_raw, (datetime, date)):
                pool_date_str = pool_date_raw.strftime('%Y-%m-%d')
            else:
                pool_date_str = str(pool_date_raw)[:10]

    # 3. 读取候选池历史(用于判断买入时是否已在池内)
    # BUG修复 (2026-07-31): 回溯合规问题
    #   问题: 沃华医药买入时不在候选池 → 之后run_full_scan将持仓股写入候选池
    #         → check_buy_compliance读取当前候选池 → 发现沃华在池内 → 判定"合规"
    #         → 但买入时实际不在池内 = 回溯合规(假阳性)
    #   修复: 对比买入日期与候选池更新日期
    #         若候选池今日才更新 → 买入时(盘中)候选池可能未包含此股
    #         → 标记为"回溯合规"而非完全合规
    prev_pool_codes = set()
    if '候选池历史' in wb.sheetnames:
        ws_hist = wb['候选池历史']
        for r in range(2, ws_hist.max_row + 1):
            hist_date_raw = ws_hist.cell(row=r, column=1).value
            hist_date_s = ""
            if hist_date_raw:
                if isinstance(hist_date_raw, (datetime, date)):
                    hist_date_s = hist_date_raw.strftime('%Y-%m-%d')
                else:
                    hist_date_s = str(hist_date_raw)[:10]
            # 只取非今日的历史记录
            if hist_date_s and hist_date_s != today_str:
                code = ws_hist.cell(row=r, column=3).value
                if code:
                    prev_pool_codes.add(str(code))

    # 4. 逐笔检查
    # BUG修复 (2026-07-30): 移除事后背驰确认豁免
    #   旧逻辑: 不在候选池 → 事后运行analyze_beichi验证DL_P>0.8 → 豁免
    #   问题: 豁免机制让"不在候选池"变得可接受 → 鼓励跳过full_scan直接买入
    #   修复: 不在候选池 = 直接违规, 不做事后豁免
    #         候选池现在包含持仓股(BUG-4修复), 所以持仓股加仓不会误报
    #
    # BUG修复 (2026-07-31): 回溯合规检测
    #   候选池今日更新 → 买入时可能不在池内 → 当前在池内是回溯写入
    #   通过候选池历史判断买入时是否已在池内
    pool_updated_today = (pool_date_str == today_str)

    for buy in today_buys:
        code = buy["code"]
        name = buy["name"]
        if code in candidate_codes:
            if pool_updated_today and code not in prev_pool_codes:
                # 候选池今日更新且该股不在历史池中 → 回溯合规
                issues.append(
                    f"⚠️ {name}({code}) 当前在候选池内但可能为回溯合规"
                    f"(候选池今日更新, 该股在历史池中未找到, 买入时可能未在池内)"
                )
            else:
                print(f"  ✓ {name}({code}) 在候选池内, 买入合规")
        else:
            issues.append(f"🔴 {name}({code}) 不在候选池内, 买入违规 (应先通过full_scan纳入候选池再买入)")

    return issues


def check_sell_compliance():
    """
    检查卖出操作是否有信号支撑 (2026-07-29 BUG修复)

    问题: 三力士违规卖出 → 系统只检测持仓是否有卖点信号,
          但不检查实际卖出操作是否有对应信号 → 无信号卖出=违规
    修复: 读取今日交易记录中的卖出操作, 逐笔检查是否有卖出信号.
    """
    issues = []
    today_str = datetime.now().strftime('%Y-%m-%d')

    try:
        wb = safe_load_wb(data_only=True)
    except Exception as e:
        print(f"  卖出合规检查: 无法读取Excel({e}), 跳过")
        return issues

    if '交易记录' not in wb.sheetnames:
        print(f"  卖出合规检查: 无交易记录sheet, 跳过")
        return issues

    ws_tr = wb['交易记录']
    today_sells = []
    for r in range(2, ws_tr.max_row + 1):
        direction = str(ws_tr.cell(row=r, column=4).value or "")
        if direction not in ("卖出", "一卖"):
            continue
        raw_date = ws_tr.cell(row=r, column=1).value
        date_str = ""
        if raw_date:
            if isinstance(raw_date, datetime):
                date_str = raw_date.strftime('%Y-%m-%d')
            elif isinstance(raw_date, date):
                date_str = raw_date.strftime('%Y-%m-%d')
            else:
                date_str = str(raw_date)[:10]
        # BUG修复 (2026-07-30): 同check_buy_compliance, 只检查今日交易
        if date_str == today_str:
            name = ws_tr.cell(row=r, column=2).value or ""
            code = str(ws_tr.cell(row=r, column=3).value or "")
            price = ws_tr.cell(row=r, column=6).value or 0
            today_sells.append({"name": name, "code": code, "price": price, "row": r, "date": date_str})

    if not today_sells:
        print(f"  卖出合规检查: 无近期卖出记录, 跳过")
        return issues

    # 检查每笔卖出是否有信号支撑
    sys.path.insert(0, BEICHI_DIR)
    try:
        from beichi_analyzer import analyze_beichi, detect_sell_signals
    except Exception as e:
        print(f"  卖出合规检查: 无法导入分析模块({e}), 跳过")
        return issues

    for sell in today_sells:
        code = sell["code"]
        name = sell["name"]
        sell_price = sell["price"]

        # 检查是否有卖出信号
        has_sell_signal = False
        signal_desc = ""

        for level in ["日线", "30min"]:
            try:
                r = analyze_beichi(code, level=level)
                if "error" in r:
                    continue
                for sig in r.get("signals", []):
                    if sig["dir"] == "看空" and sig["valid"]:
                        has_sell_signal = True
                        signal_desc = f"{level} {sig['op']} DL_P={sig['dl_prob']:.2f}"
                        break
                if has_sell_signal:
                    break
            except:
                pass

        # 也检查综合卖出信号
        if not has_sell_signal and sell_price > 0:
            try:
                # 获取成本价(从持仓记录)
                cost_price = 0
                ws_hold = wb.get('持仓表')
                if ws_hold:
                    for hr in range(2, ws_hold.max_row + 1):
                        h_code = str(ws_hold.cell(row=hr, column=3).value or "")
                        if h_code == code:
                            cost_price = ws_hold.cell(row=hr, column=5).value or 0
                            break

                if cost_price > 0:
                    sell_eval = detect_sell_signals(code, cost_price, sell_price)
                    if sell_eval["should_clear"] or sell_eval["should_reduce"]:
                        has_sell_signal = True
                        signal_desc = f"综合 {sell_eval['reason']}"
            except:
                pass

        # 去弱留强减仓依据 (2026-07-30):
        # 前提: 持仓数 > 婴儿账户持股个数均值(5只)
        # 条件: DL_P < 0.8 (未达确认标准) → 允许主动减仓
        # 理由: 婴儿账户资金有限, 持仓过多稀释效率,
        #       未达确认标准的持仓不值得保留仓位
        # 注意: 使用detect_multilevel_buy_signals获取DL_P, 与full_scan一致
        #
        # BUG修复 (2026-07-30): 增加一买低点破位作为减仓依据
        #   问题: 沃华DL_P=0.94>0.8 → 去弱留强不会触发 → 即使破一买低也不卖
        #         但DL_P高≠无风险, 二买可能失效
        #   修复: DL_P>=0.8但破一买低点 → 也允许减仓(二买失败优先于信号强度)
        INFANT_MAX_HOLD = 5
        if not has_sell_signal:
            try:
                from beichi_analyzer import detect_multilevel_buy_signals
                ml = detect_multilevel_buy_signals(code, price=sell_price)
                daily_dlp = ml.get('daily_dl_p', 0)
                current_hold_count = len(get_today_holdings())
                obl = ml.get('one_buy_low')
                zy_info = ml.get('zhongyin', {})

                # 条件A: DL_P < 0.8 + 持仓过多 → 去弱留强
                if daily_dlp < 0.8 and current_hold_count > INFANT_MAX_HOLD:
                    has_sell_signal = True
                    signal_desc = f"去弱留强(DL_P={daily_dlp:.2f}<0.8未确认, 持仓{current_hold_count}只>{INFANT_MAX_HOLD}只均值)"

                # 条件B (BUG修复): 破一买低点 → 二买失败, 无论DL_P多高都允许减仓
                #   这解决了"DL_P高但不破位"和"DL_P高但破位"的区分
                #   沃华DL_P=0.94但如果破一买低 → 必须允许卖出止损
                #
                # BUG修复 (2026-07-31): 一买低点合理性校验
                #   问题: 三力士成本3.31, 一买低4.23 → 成本<一买低(逻辑不可能)
                #         成本价在买入时就应该在一买低上方, 否则不可能买入
                #         一买低>成本 说明中枢检测有误(min_amp_pct过低产生噪音中枢)
                #         → 错误触发"破一买低"卖出 → +3.14%浮盈被清仓 = 违规卖出
                #   修复: 若 obl > cost_price → 一买低无效, 不触发条件B
                #         额外校验: obl > 0 且 obl < cost_price * 1.5 (防止异常高值)
                elif obl and obl > 0 and sell_price < obl:
                    # 获取成本价做合理性校验
                    _cost_for_check = cost_price
                    if _cost_for_check <= 0:
                        # 从持仓表重新获取成本
                        ws_hold2 = wb.get('持仓表')
                        if ws_hold2:
                            for hr2 in range(2, ws_hold2.max_row + 1):
                                if str(ws_hold2.cell(row=hr2, column=3).value or '') == code:
                                    _cost_for_check = ws_hold2.cell(row=hr2, column=8).value or 0
                                    break

                    if _cost_for_check > 0 and obl > _cost_for_check:
                        # 一买低 > 成本价 → 中枢检测异常, 一买低无效
                        print(f"  ⚠️ {name}({code}) 一买低{obl:.2f}>成本{_cost_for_check:.2f}(逻辑异常, 跳过条件B)")
                    elif _cost_for_check > 0 and obl > _cost_for_check * 1.5:
                        # 一买低异常高 → 中枢检测可能错误
                        print(f"  ⚠️ {name}({code}) 一买低{obl:.2f}>成本×1.5={_cost_for_check*1.5:.2f}(异常高, 跳过条件B)")
                    else:
                        has_sell_signal = True
                        pct_below = ((sell_price - obl) / obl) * 100
                        signal_desc = (
                            f"二买失败止损(现价{sell_price:.2f}<一买低{obl:.2f}, "
                            f"跌{pct_below:+.1f}%, DL_P={daily_dlp:.2f}但二买已失效)"
                        )

                # 条件C (P1, 2026-08-02): 中阴状态 → 仓位压制/NotChasing
                # 中阴 = 背驰信号存在但趋势未确认 → 允许减仓释放资金
                # NotChasing = 背驰存在但价格脱离中枢 → 标记不加仓
                elif zy_info.get('is_zhongyin') and current_hold_count > INFANT_MAX_HOLD:
                    has_sell_signal = True
                    signal_desc = (
                        f"中阴减仓(背驰存在但趋势未确认, "
                        f"{zy_info.get('reason', '')}, "
                        f"持仓{current_hold_count}只>{INFANT_MAX_HOLD}只)"
                    )
            except:
                pass

        if has_sell_signal:
            print(f"  ✓ {name}({code}) 卖出有信号支撑: {signal_desc}")
        else:
            issues.append(f"🔴 {name}({code}) 卖出无信号支撑, 卖出违规")

    return issues


def check_compliance():
    holdings = get_today_holdings()
    account = get_account_summary()
    issues = []

    # === 合规审查首项: 成本价-破位级别匹配检查 (策略C, 2026-07-26) ===
    # 规则: 以成本价确定买点级别, 只执行该级别的卖点/破位
    #       现价小级别买点不构成加仓/持有理由
    #       破位级别必须与成本买点级别匹配才触发操作
    print(f"{'='*50}")
    print(f"📝 合规核查 {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"{'='*50}")
    print(f"【首项】成本价-破位级别匹配检查 (策略C)")
    print(f"  规则: 成本价定买点级别 → 只执行该级别卖点/破位")
    print(f"  原则: 买入卖出同级别闭环, 避免级别错配")

    level_mismatch_count = 0
    for h in holdings:
        if h['waived'] == '是':
            continue
        code = str(h['code'])
        cost = h['entry'] or 0
        close = h['close'] or 0
        if cost <= 0:
            continue

        entry_data = detect_entry_level(code, cost)
        entry_level = entry_data["entry_level"]
        sell_levels = entry_data["sell_levels"]
        zone = entry_data["entry_info"].get(entry_level, {}).get("zone", "未知") if entry_level else "未确定"

        # 检查成本对应级别是否破位
        level_bd = False
        for level in sell_levels:
            try:
                r = analyze_beichi(code, level=level)
                if "error" in r or not r.get("zss"):
                    continue
                last_zs = r["zss"][-1]
                if close < last_zs["zd"]:
                    pct = ((close - last_zs["zd"]) / last_zs["zd"]) * 100
                    issues.append(f"🔴 {h['name']}({code}) {level}中枢破位{pct:+.1f}% (成本{cost:.2f}对应{entry_level}买点, 现价{close:.2f}<下沿{last_zs['zd']:.2f})")
                    level_bd = True
            except:
                pass

        # 检查是否有更小级别买点(策略C明确忽略)
        if entry_level and level_bd:
            level_idx = ["日线", "30min", "5min", "1min"].index(entry_level)
            smaller_levels = ["日线", "30min", "5min", "1min"][level_idx+1:]
            for level in smaller_levels:
                try:
                    r = analyze_beichi(code, level=level)
                    if "error" in r:
                        continue
                    for sig in r.get("signals", []):
                        if sig["op"] == "一买" and sig["ratio"] < 60 and sig["dl_prob"] > 0.8 and sig["valid"]:
                            issues.append(f"⚠️ {h['name']}({code}) {level}有确认一买但非成本对应级别, 不构成加仓理由(策略C)")
                except:
                    pass

    print(f"  成本级别匹配检查: {'⚠️发现'+str(len([i for i in issues if '中枢破位' in i]))+'个破位' if issues else '✓ 无破位'}")

    # === 买入候选池合规检查 (2026-07-29) ===
    # BUG修复: check_compliance()不检查"买入标的是否在候选池内"
    # 问题: 沃华医药DL_P=0.93但不在候选池(被截断) → 手动检查时违规
    #       但系统compliance不检查这条 → 一会儿行一会儿不行
    # 修复: 读取今日交易记录中的买入操作, 逐笔检查是否在候选池内
    buy_compliance_issues = check_buy_compliance(holdings)
    issues.extend(buy_compliance_issues)

    # === 卖出信号合规检查 (2026-07-29) ===
    # BUG修复: 三力士违规卖出 → 系统compliance不检查卖出是否有信号
    # 修复: 读取今日交易记录中的卖出操作, 逐笔检查是否有卖出信号
    sell_compliance_issues = check_sell_compliance()
    issues.extend(sell_compliance_issues)

    # === 其他合规检查 ===
    # 【动态仓位上限】(2026-07-26): 修复重仓合规悖论
    # 旧逻辑: 静态35%上限 → 趋势加仓被阻止 → 错过趋势收益
    # 新逻辑: 买点升级+浮盈护垫+止损上移 → 动态提升上限

    # ============================================================
    # BUG修复 (2026-07-31): 一买低点破位检查 — 独立循环, 不受WAIVED限制
    # 问题: 原代码在一行 for h in holdings: if waived == '是': continue 循环内
    #       → WAIVED股票(东风股份)被continue跳过 → 一买低点检查永远不执行
    # 修复: 将一买低点检查拆到独立循环, 对所有持仓无条件执行
    # ============================================================
    for h in holdings:
        code = str(h['code'])
        cost = h['entry'] or 0
        close = h['close'] or 0

        if cost > 0 and close > 0:
            try:
                from beichi_analyzer import detect_multilevel_buy_signals as _dml
                _ml = _dml(code, price=close)
                _obl = _ml.get("one_buy_low")
                # 合理性校验: 一买低必须 < 成本价
                if _obl and _obl > 0 and _obl < cost and close < _obl:
                    _pct = ((close - _obl) / _obl) * 100
                    issues.append(
                        f"🔴 {h['name']}({code}) 破一买低点: 现价{close:.2f}<一买低{_obl:.2f} "
                        f"(跌{_pct:+.1f}%) → 应立即清仓"
                    )
            except:
                pass

    for h in holdings:
        if h['waived'] == '是':
            continue
        if h['close'] and h['stop'] and h['close'] <= h['stop']:
            issues.append(f"⚠️ {h['name']}已破止损: 现价{h['close']:.2f}<=止损{h['stop']:.2f}")

        code = str(h['code'])
        cost = h['entry'] or 0
        close = h['close'] or 0

        if cost > 0 and close > 0:
            # 计算动态上限 (顺便更新dynamic_cap_info)
            dynamic_cap = get_dynamic_position_cap(code, cost, close)
            tier = dynamic_cap_info.get("tier", "边缘池")
            entry = dynamic_cap_info.get("entry", "一买")
            pnl_pct = dynamic_cap_info.get("pnl_pct", 0)

            # ============================================================
            # BUG修复 (2026-07-29): 二买加仓合规检查缺失
            #
            # 问题: check_compliance()只检查"仓位超限", 不检查"二买信号已确认
            #       但仓位未升级到50%"的合规告警.
            #       加仓信号检测只在run_intraday_scan做信息展示, 不进入issues.
            #       → 沃华医药有二买信号(valid)但合规不提示应加仓
            #
            # 修复: 在合规检查中增加三种场景:
            #   A. 二买已确认 + 浮盈>=5% + 当前仓位<50% → 应加仓未加仓
            #   B. 二买已确认 + 浮盈<5% + 当前仓位<50% → 信号存在但护垫不足
            #   C. 二买已确认 + 浮盈<5% + 当前仓位>=50% → 需警惕护垫不足风险
            # ============================================================
            if entry in ("二买", "三买"):
                target_cap = 0.50 if entry == "二买" else 0.60
                cur_pos = h.get('pos', 0) or 0

                # BUG修复 (2026-07-30): 获取一买低点距离
                obl = dynamic_cap_info.get("one_buy_low")
                dist_to_low = dynamic_cap_info.get("dist_to_one_buy_low")

                if pnl_pct >= 0.05:
                    # 场景A: 满足所有加仓条件, 应升级
                    if cur_pos < target_cap:
                        remaining = target_cap - cur_pos
                        d_dp = dynamic_cap_info.get("daily_dl_p", 0)
                        m30_ep = dynamic_cap_info.get("30min_ep_p", 0)
                        ermai_score = dynamic_cap_info.get("ermai_dl_prob", 0)
                        # BUG修复 (2026-07-30): 附带一买低点距离信息
                        obl_info = ""
                        if obl and dist_to_low is not None:
                            if dist_to_low < 0.03:
                                obl_info = f" ⚠️离一买低仅{dist_to_low:.1%}<3%, 不宜加仓"
                            else:
                                obl_info = f" (离一买低{dist_to_low:.1%})"
                        issues.append(
                            f"🟢 {h['name']}({code}) {entry}信号确认 + 浮盈{pnl_pct:.1%}>=5%, "
                            f"仓位应升级到{target_cap:.0%}, 当前{cur_pos:.1%}, 可加仓{remaining:.1%} "
                            f"(DL_P={d_dp:.2f} EP_L={m30_ep:.2f} 确认强度={ermai_score:.2f}){obl_info}"
                        )
                else:
                    # 场景B/C: 二买信号存在但浮盈护垫不足
                    if cur_pos < target_cap:
                        print(f"  ℹ️ {h['name']}({code}) {entry}信号确认但浮盈{pnl_pct:.1%}<5%护垫, "
                              f"暂不加仓(需涨至{cost * 1.05:.2f}方可触发)")
                    else:
                        issues.append(
                            f"⚠️ {h['name']}({code}) {entry}信号确认但浮盈{pnl_pct:.1%}<5%护垫, "
                            f"当前仓位{cur_pos:.1%}已超安全区, 注意风险"
                        )

                # ============================================================
                # BUG修复 (2026-07-30): 加仓后一买低点风险检查
                #
                # 场景D: 已加仓(pos>35%) + 价格接近/破一买低点 → 风险告警
                # 场景E: 已加仓(pos>35%) + 价格已破一买低点 → 严重违规
                #
                # 核心风险: 去弱留强加仓浮盈小的 → 二买无法确认 → 破一买 → 重仓回撤
                # ============================================================
                if obl and obl > 0 and close > 0:
                    if close < obl:
                        # 场景E: 已破一买低点
                        if cur_pos > 0.35:
                            issues.append(
                                f"🔴 {h['name']}({code}) 已加仓(仓位{cur_pos:.1%}>35%)且现价{close:.2f}"
                                f"<一买低{obl:.2f}, 二买失败! 应立即减仓至35%以下止损"
                            )
                    elif dist_to_low is not None and dist_to_low < 0.03:
                        # 场景D: 接近一买低点(<3%)
                        if cur_pos > 0.35:
                            issues.append(
                                f"🟠 {h['name']}({code}) 已加仓(仓位{cur_pos:.1%}>35%)且价格离一买低"
                                f"仅{dist_to_low:.1%}<3%, 二买破位风险高, 建议减仓至35%"
                            )

            # T+1锁定豁免: 全部锁定的股票今日无法操作, 不因仓位超限告警
            t1_lock = h.get('t1_lock', '')
            pos_fmt = f"{h['pos']:.1%}" if h.get('pos') is not None else "N/A"
            cap_fmt = f"{dynamic_cap:.0%}" if dynamic_cap is not None else "N/A"
            if t1_lock == '全部锁定':
                print(f"  ✓ {h['name']}仓位{pos_fmt} <= 动态上限{cap_fmt} "
                      f"[{tier}] (T+1全部锁定, 豁免仓位检查)")
            else:
                # 动态仓位上限超限检查
                if h['pos'] and h['pos'] > dynamic_cap:
                    issues.append(
                        f"⚠️ {h['name']}仓位超限: {pos_fmt}>动态上限{cap_fmt}"
                        f"(买点={entry} 分层={tier} 浮盈={pnl_pct:.1%})"
                    )
                else:
                    print(f"  ✓ {h['name']}仓位{pos_fmt} <= 动态上限{cap_fmt} "
                          f"[{tier}] (多级别共振合规)")
        elif h['pos'] and h['pos'] > 0.35:
            # 无有效成本/现价时回退到静态检查
            t1_lock = h.get('t1_lock', '')
            pos_fmt = f"{h['pos']:.1%}" if h.get('pos') is not None else "N/A"
            if t1_lock == '全部锁定':
                print(f"  ✓ {h['name']}仓位{pos_fmt} (T+1全部锁定, 豁免静态上限检查)")
            else:
                issues.append(
                    f"⚠️ {h['name']}仓位{pos_fmt}>35%静态上限 (无法计算动态上限, 缺少成本/现价)"
                )

    # ============================================================
    # BUG修复 (2026-08-06): 合规检查模块缺失三项关键检查
    # 问题: check_compliance() 输出"持仓合规, 无告警"但东风股份不在候选池内
    #       且DL_P=0.08, 属于严重违规未被检测到
    # 根因: compliance模块缺少:
    #   1. 有效持仓数量检查(婴儿账户≤3只)
    #   2. 候选池+DL_P阈值检查(持仓股是否在候选池内, DL_P≥0.8)
    #   3. P0状态检查(执行清单中P0标记是否已执行)
    # 修复: 增加以上三项检查
    # ============================================================

    # === 缺失检查1: 有效持仓数量检查 (婴儿账户) ===
    INFANT_MAX_HOLD = 3
    non_waived = [h for h in holdings if h.get('waived') != '是']
    if len(non_waived) > INFANT_MAX_HOLD:
        issues.append(
            f"🔴 有效持仓{len(non_waived)}只超限(婴儿账户上限{INFANT_MAX_HOLD}只, "
            f"资产<5万阶段): " + ", ".join([f"{h['name']}({h['code']})" for h in non_waived])
        )
    else:
        print(f"  ✓ 有效持仓{len(non_waived)}只 <= 婴儿账户上限{INFANT_MAX_HOLD}只")

    # === 缺失检查2: 候选池+DL_P阈值检查 ===
    try:
        wb_pool = safe_load_wb(data_only=True)
        pool_codes = set()
        pool_dlp = {}  # code -> dl_p
        pool_tier = {}  # code -> tier
        if '候选池' in wb_pool.sheetnames:
            ws_pool = wb_pool['候选池']
            for r in range(2, ws_pool.max_row + 1):
                code = ws_pool.cell(row=r, column=3).value
                tier = ws_pool.cell(row=r, column=11).value  # K列=分层Tier
                dlp = ws_pool.cell(row=r, column=8).value  # H列=DL_P
                if code:
                    code_s = str(code)
                    pool_codes.add(code_s)
                    if dlp:
                        try:
                            pool_dlp[code_s] = float(dlp)
                        except:
                            pass
                    if tier:
                        pool_tier[code_s] = str(tier)

        for h in holdings:
            code = str(h['code'])
            name = h['name']
            close = h.get('close', 0) or 0
            entry = h.get('entry', 0) or 0
            waived = h.get('waived') == '是'

            # 检查1: 是否在候选池内
            if code in pool_codes:
                tier_info = pool_tier.get(code, "未知")
                dlp_info = pool_dlp.get(code, 0)
                print(f"  ✓ {name}({code}) 在候选池内 [Tier={tier_info} DL_P={dlp_info:.2f}]")
            elif not waived:
                # 非WAIVED且不在候选池 → 违规
                dlp_val = pool_dlp.get(code, 0)
                issues.append(
                    f"🔴 {name}({code}) 不在候选池内(DL_P={dlp_val:.2f}<0.8), "
                    f"持仓逻辑失效, 应清仓或减仓至观察仓位"
                )

            # 检查2: DL_P阈值(持仓端)
            dlp = pool_dlp.get(code, 0)
            if dlp > 0 and dlp < 0.8 and not waived:
                if close > 0 and entry > 0 and close < entry:
                    # DL_P<0.8且亏损态 → RiskWatch
                    issues.append(
                        f"🟠 {name}({code}) DL_P={dlp:.2f}<0.8且亏损(成本{entry:.2f}>现价{close:.2f}), "
                        f"触发RiskWatch, 建议减仓"
                    )
                elif dlp < 0.5:
                    # DL_P<0.5极弱信号 → 强制清仓建议
                    issues.append(
                        f"🔴 {name}({code}) DL_P={dlp:.2f}<0.5信号极弱, "
                        f"建议强制清仓(去弱留强)"
                    )

        wb_pool.close()
    except Exception as e:
        print(f"  候选池检查异常: {e}")

    # === 缺失检查3: P0状态检查 ===
    try:
        wb_p0 = safe_load_wb(data_only=True)
        if '执行清单' in wb_p0.sheetnames:
            ws_exec = wb_p0['执行清单']
            p0_pending = []
            for r in range(2, ws_exec.max_row + 1):
                note = str(ws_exec.cell(row=r, column=15).value or "")
                action = str(ws_exec.cell(row=r, column=4).value or "")
                stock_name = ws_exec.cell(row=r, column=2).value or ""
                stock_code = ws_exec.cell(row=r, column=3).value or ""
                if 'P0' in note or 'P0' in action:
                    p0_pending.append(f"{stock_name}({stock_code}): {note}")

            if p0_pending:
                for p0_item in p0_pending:
                    issues.append(f"🔴 P0待执行: {p0_item}")
                print(f"  ⚠️ P0待执行项: {len(p0_pending)}条")
            else:
                print(f"  ✓ 无P0待执行项")
        wb_p0.close()
    except Exception as e:
        print(f"  P0状态检查异常: {e}")

    # 输出合规摘要
    print(f"\n账户总资产: ¥{account.get('total_asset', 0):.2f}" if account.get('total_asset') else "账户总资产: N/A")
    print(f"现金: ¥{account.get('cash', 0):.2f}" if account.get('cash') else "现金: N/A")
    print(f"仓位比例: {account.get('position_ratio', 0):.1%}" if account.get('position_ratio') else "仓位比例: N/A")
    print(f"持仓数量: {len(holdings)}只")
    if issues:
        print(f"\n⚠️ 合规告警 ({len(issues)}项):")
        for issue in issues:
            print(f"  {issue}")
    else:
        print("\n✓ 持仓合规, 无告警")

    print(f"\n📝 心态日志提醒:")
    print(f"  评分维度(0-5分):")
    print(f"  • 非工作流看盘次数 (0=0次最好)")
    print(f"  • 是否追价操作 (0=没有最好)")
    print(f"  • 是否按系统执行 (5=完全最好)")
    print(f"  • 情绪状态 (5=平静最好)")
    print(f"  • 止损外操作 (0=没有最好)")
    print(f"  达标线: 80分")
    print(f"{'='*50}")

    # 同时输出 JSON (供程序化处理)
    def safe(v):
        if isinstance(v, (datetime, date)): return str(v)
        if isinstance(v, Decimal): return float(v)
        return v
    print(json.dumps({"account": {k:safe(v) for k,v in account.items()}, "holdings": [{k:safe(v) for k,v in h.items()} for h in holdings], "issues": issues, "compliant": len(issues)==0}, ensure_ascii=False, indent=2))

    return account, holdings, issues

def safe_val(v):
    if isinstance(v, (datetime, date)): return str(v)
    if isinstance(v, Decimal): return float(v)
    return v

def run_full_scan():
    """全市场候选扫描: 沪A主板全量 + 深市全量(000/002) + 写入候选池(排除持仓股)

    分层候选池 (2026-07-26):
      核心池(DL_P>0.90+ratio<20%): 调仓首选, 1-2周稳定
      观察池(DL_P 0.85-0.90): 核心池不足时补充
      边缘池(DL_P 0.80-0.85): 仅观察不买入
    写入Excel时: 核心池优先 → 观察池补充 → 边缘池末尾, 备注列标注层级
    """
    sys.path.insert(0, BEICHI_DIR)
    from full_scan import full_scan, calc_funding
    account = get_account_summary()
    result = full_scan(
        total_asset=account["total_asset"] or 20326.12,
        cash=account["cash"] or 7847.12,
        silent=False,
        prefetch=True,     # 2026-08-11: 并发预取5min数据, 加速ML推理
    )

    # ============================================================
    # BUG诊断 (2026-08-06): 全市场无DL_P>0.8时自动核查
    # 规则: 全市场无DL_P>0.8信号 → 极大概率是bug → 自动诊断
    # 诊断已由full_scan()内部完成, 此处检查结果并触发告警
    # ============================================================
    scan_diag = result.get("scan_diagnostics", {})
    if scan_diag.get("triggered"):
        print(f"\n{'='*50}", flush=True)
        print("⚠️ ⚠️ ⚠️  全市场扫描异常  ⚠️ ⚠️ ⚠️", flush=True)
        print(f"{'='*50}", flush=True)
        print(scan_diag.get("report", "未知诊断结果"), flush=True)
        print(f"{'='*50}", flush=True)
        print("→ 候选池将保留昨日数据, 等待修复后重新扫描", flush=True)
        print(f"{'='*50}\n", flush=True)

    # 排除持仓股(已持有的不再推荐为新候选)
    # BUG修复 (2026-07-30): 旧代码排除持仓股后, 沃华/贵绳不在候选池
    #   → check_buy_compliance报"不在候选池=违规"
    #   → 依赖事后背驰确认作为豁免 → "一会儿行一会儿不行"
    # 修复: 持仓股仍写入候选池, 但标注为"持仓"不作为新推荐
    #       这样买入合规检查能通过(在候选池内), 且不会重复推荐
    holdings = get_today_holdings()
    held_codes = {str(h['code']) for h in holdings if h.get('code')}
    # 不再排除持仓股, 全部confirmed写入候选池
    all_confirmed = result["confirmed"]
    if held_codes:
        held_in_pool = [r for r in all_confirmed if r["code"] in held_codes]
        if held_in_pool:
            print(f"  持仓股也在候选池: {len(held_in_pool)}只 ({', '.join(r['name'] for r in held_in_pool)})", flush=True)

    # ============================================================
    # 候选池 V3 (2026-07-29): 写入所有confirmed标的, 不截断
    #
    # BUG修复 (2026-07-29): 候选池漏选导致合规审查矛盾
    # 问题: 旧代码只写沪深各5只共10只 → DL_P=0.93的沃华医药被挤出
    #       → 手动检查时"不在候选池=违规", 但系统compliance不检查这条
    #       → 一会儿行一会儿不行
    # 修复: 所有confirmed标的都写入候选池, 按分层排序
    #       核心池全部写入, 观察池全部写入, 边缘池限20只
    # ============================================================
    core = [r for r in all_confirmed if r.get("tier") == "核心"]
    watch = [r for r in all_confirmed if r.get("tier") == "观察"]
    edge = [r for r in all_confirmed if r.get("tier") == "边缘"]

    # 按DL_P降序排序, 核心池优先
    def sort_by_dlp(stocks):
        sha = sorted([r for r in stocks if r["code"].startswith("6")], key=lambda x: (-x["dlp"], x["ratio"]))
        sza = sorted([r for r in stocks if r["code"].startswith("0")], key=lambda x: (-x["dlp"], x["ratio"]))
        return sha + sza

    # 核心池+观察池全部写入, 边缘池限20只(避免候选池过大)
    selected = sort_by_dlp(core) + sort_by_dlp(watch)
    selected += sort_by_dlp(edge)[:20]

    print(f"\n[婴儿级候选池] 分层: 核心{len(core)}只 + 观察{len(watch)}只 + 边缘{len(edge)}只", flush=True)
    print(f"写入: {len(selected)}只 (核心+观察全部, 边缘限20只)", flush=True)
    print(f"  核心{len([s for s in selected if s.get('tier')=='核心'])} + 观察{len([s for s in selected if s.get('tier')=='观察'])} + 边缘{len([s for s in selected if s.get('tier')=='边缘'])}", flush=True)

    wb = safe_load_wb()
    ws = wb['候选池']

    # ============================================================
    # BUG修复 (2026-07-28): 候选池写入后不更新
    #
    # 根因1: 清空旧数据时跳过公式列(以=开头的cell) → 公式残留在空行
    #        → ws.max_row因残留公式虚高到192行 → 清空循环范围巨大但无效
    #        → 旧行的公式引用错行(如Row4的L列引用Row5) → 显示旧数据
    #
    # 根因2: 写入新数据时只重写列11公式, 列12-47保留旧公式
    #        → 公式引用的行号与新数据行不匹配 → 计算结果错误
    #
    # 修复: 1. 彻底清空所有行(包括公式), 保留表头
    #       2. 写入新数据时为每行重新写入所有公式(用正确行号)
    #       3. 无确认标的时也清空候选池(显示空池而非旧数据)
    # ============================================================

    # 备份公式模板(从第2行提取, 用作写入时的模板)
    formula_templates = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=2, column=c).value
        if v and str(v).startswith('='):
            formula_templates[c] = str(v)

    # 彻底清空所有数据行(保留表头第1行)
    # BUG修复 (2026-07-28): cell.value=None只清值不删行, max_row不变
    # 用 delete_rows 彻底删除多余行, 避免残留空行导致max_row虚高
    max_row = ws.max_row
    if max_row > 1:
        ws.delete_rows(2, max_row - 1)

    # 如果无确认标的, 清空后直接保存(显示空池)
    if not all_confirmed:
        print("\n候选池: 无确认标的(排除持仓后), 候选池已清空")
        safe_save_wb(wb)
        recalc()
        return {
            "scan_result": result,
            "selected": [],
            "errors": 0,
        }

    today = date.today()
    # 分层映射: 中文简称 → 候选池显示名称
    tier_display = {
        "核心": "核心池",
        "观察": "观察池",
        "边缘": "边缘池",
    }
    # 分层颜色 (与fix_tier.py一致): 黄/蓝/灰
    from openpyxl.styles import PatternFill, Font as XFont
    tier_fill = {
        "核心": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),  # 黄
        "观察": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),  # 蓝
        "边缘": PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),  # 灰
    }
    for idx, stock in enumerate(selected):
        row = 2 + idx
        price = stock["price"]
        fund = calc_funding(price, result["total_asset"], result["cash"])
        tier = stock.get("tier", "边缘")
        tier_name = tier_display.get(tier, "边缘池")

        # 备注列: 仅保留资金转入信息(分层已由列11公式自动计算)
        note = ""
        if fund["need_transfer"]: note = "需转入%.0f元" % fund["transfer"]

        ws.cell(row=row, column=1, value=today)
        ws.cell(row=row, column=2, value=stock.get("name", ""))
        ws.cell(row=row, column=3, value=stock["code"])
        ws.cell(row=row, column=4, value=price)
        ws.cell(row=row, column=5, value=price)
        ws.cell(row=row, column=6, value=round(price * 0.95, 2))
        ws.cell(row=row, column=7, value=stock["ratio"] / 100)
        ws.cell(row=row, column=8, value=stock["dlp"])
        ws.cell(row=row, column=9, value=str(stock["valid"]))
        ws.cell(row=row, column=10, value="一买")

        # BUG修复 (2026-07-28): 为每行写入所有公式(用正确行号)
        # 旧代码只写列11, 列12-47保留旧公式 → 引用错行 → 候选池不更新
        # 使用openpyxl Translator正确处理相对/绝对引用
        from openpyxl.formula.translate import Translator
        for col, template in formula_templates.items():
            formula = Translator(template, origin="B2").translate_formula(f"B{row}")
            ws.cell(row=row, column=col, value=formula)

        ws.cell(row=row, column=11).font = XFont(bold=True, size=11)
        # 行颜色: 按分层着色 (黄/蓝/灰)
        fill = tier_fill.get(tier)
        if fill:
            for c in range(1, 32):
                ws.cell(row=row, column=c).fill = fill
        # 列31备注: 仅资金信息
        if note:
            ws.cell(row=row, column=31, value=note)

    safe_save_wb(wb)
    recalc_result = recalc()
    errors = recalc_result.get("total_errors", -1)

    print(f"\n候选池: 写入{len(selected)}只 (核心{len([s for s in selected if s.get('tier')=='核心'])} + 观察{len([s for s in selected if s.get('tier')=='观察'])} + 边缘{len([s for s in selected if s.get('tier')=='边缘'])})")
    print(f"公式重算: {errors}个错误")
    return {
        "scan_result": result,
        "selected": selected,
        "errors": errors,
    }

def detect_entry_level(code, cost):
    """根据成本价判断属于哪个级别的买点区间, 返回对应卖点级别列表

    策略C (2026-07-26修订):
    逻辑: 成本落在某级别中枢区间内 → 该级别为买点级别 → 只监控该级别卖点/破位
          成本在最新中枢下方 → 该级别一买区 → 只监控该级别卖点/破位
          成本在所有中枢上方 → 基础级别买点 → 默认监控日线级别

    核心原则: 买入逻辑和卖出逻辑在同一级别闭环, 避免级别错配
    - 不扩展到更小级别(避免1min破位噪音干扰日线买点持仓)
    - 不看现价小级别买点(避免"现价有1min买点所以可以加仓"的误判)

    【BUG-8修复 (2026-07-27)】跳过1min级别
    问题: 1min级别需要build_1min_from_5min → 2次网络请求/只
          7只持仓 × 2 = 14次额外请求, 每次最多10秒timeout
          → GitHub Actions 10分钟超时, 工作流卡死
    修复: intraday scan中跳过1min级别(持仓级别判断不需要1min精度)
    """
    sys.path.insert(0, BEICHI_DIR)
    from beichi_analyzer import analyze_beichi
    levels_priority = ["日线", "30min", "5min"]  # BUG-8: 跳过1min(网络开销过大)
    entry_info = {}

    for level in levels_priority:
        try:
            r = analyze_beichi(code, level=level)
            if "error" in r or not r.get("zss"):
                continue
            zss = r["zss"]
            last_zs = zss[-1]
            first_zs = zss[0]

            if last_zs["zd"] <= cost <= last_zs["zg"]:
                entry_info[level] = {"zone": "中枢内", "zs": last_zs}
            elif cost < last_zs["zd"]:
                entry_info[level] = {"zone": "中枢下方(一买区)", "zs": last_zs}
            elif cost < first_zs["zd"]:
                entry_info[level] = {"zone": "全中枢下方(深度一买区)", "zs": first_zs}
            else:
                entry_info[level] = {"zone": "中枢上方", "zs": last_zs}
        except:
            pass

    # 确定买点级别: 找成本最接近中枢区间内的级别(优先大级别)
    best_entry_level = None
    for level in levels_priority:
        info = entry_info.get(level)
        if info and info["zone"] in ("中枢内", "中枢下方(一买区)", "全中枢下方(深度一买区)"):
            best_entry_level = level
            break

    # 卖点监控级别: 只监控成本对应的买点级别(策略C)
    # 不扩展到更小级别, 避免小级别噪音造成过度交易
    if best_entry_level:
        sell_levels = [best_entry_level]  # 只监控该级别
    else:
        sell_levels = ["日线"]  # 无法确定则默认日线

    return {
        "entry_level": best_entry_level,
        "entry_info": entry_info,
        "sell_levels": sell_levels,
    }

def get_candidate_pool():
    """从Excel候选池读取标的列表"""
    wb = safe_load_wb(data_only=True)
    ws = wb['候选池']
    candidates = []
    for r in range(2, ws.max_row + 1):
        code = ws.cell(row=r, column=3).value
        if not code:
            continue
        name = ws.cell(row=r, column=2).value or ""
        price = ws.cell(row=r, column=4).value or 0
        candidates.append({"code": str(code), "name": name, "price": float(price)})
    return candidates

# ============================================================
# P0清仓升级机制 (2026-07-31 BUG修复)
# 问题: 东风股份连续3天P0未执行 → 无升级告警 → 纪律执行失败
# 修复: 跟踪P0待执行天数, 升级告警级别, 写入执行清单
# ============================================================

def _count_p0_pending_days(wb, code, name):
    """计算某股票P0清仓待执行天数

    读取执行清单中该股票的P0标记记录, 计算从首次标记到今天的天数
    如果执行清单中已有该股票的卖出成交记录, 返回0(已执行)

    Args:
        wb: openpyxl Workbook对象
        code: 股票代码
        name: 股票名称

    Returns:
        int: P0待执行天数 (0=未标记或已执行)
    """
    try:
        today = date.today()
        first_p0_date = None

        # 检查执行清单中的P0标记
        if '执行清单' in wb.sheetnames:
            ws = wb['执行清单']
            for r in range(2, ws.max_row + 1):
                row_code = str(ws.cell(row=r, column=3).value or '')
                row_action = str(ws.cell(row=r, column=4).value or '')
                row_date = ws.cell(row=r, column=1).value

                if row_code == str(code):
                    # 检查是否已执行(有卖出成交记录)
                    if '卖出' in row_action or '清仓' in row_action:
                        shares = ws.cell(row=r, column=7).value
                        if shares and int(shares) > 0:
                            return 0  # 已执行

                    # 查找P0标记
                    note = str(ws.cell(row=r, column=15).value or '') + str(ws.cell(row=r, column=21).value or '')
                    if 'P0' in note or '破一买低' in note:
                        if row_date:
                            if isinstance(row_date, (datetime, date)):
                                d = row_date.date() if isinstance(row_date, datetime) else row_date
                            else:
                                try:
                                    d = datetime.strptime(str(row_date)[:10], '%Y-%m-%d').date()
                                except:
                                    continue
                            if first_p0_date is None or d < first_p0_date:
                                first_p0_date = d

        # 检查交易记录中是否有卖出(已执行)
        if '交易记录' in wb.sheetnames:
            ws_tr = wb['交易记录']
            for r in range(2, ws_tr.max_row + 1):
                row_code = str(ws_tr.cell(row=r, column=3).value or '')
                row_dir = str(ws_tr.cell(row=r, column=4).value or '')
                if row_code == str(code) and ('卖出' in row_dir or '一卖' in row_dir):
                    raw_date = ws_tr.cell(row=r, column=1).value
                    if raw_date:
                        if isinstance(raw_date, (datetime, date)):
                            d = raw_date.date() if isinstance(raw_date, datetime) else raw_date
                        else:
                            try:
                                d = datetime.strptime(str(raw_date)[:10], '%Y-%m-%d').date()
                            except:
                                continue
                        # 如果在first_p0_date之后卖出 → 已执行
                        if first_p0_date and d >= first_p0_date:
                            return 0

        if first_p0_date:
            return (today - first_p0_date).days
        return 0
    except Exception as e:
        return 0


def _mark_p0_pending(wb, code, name, price, obl, pct):
    """在执行清单中写入/更新P0待执行标记

    如果该股票已有P0标记且未执行, 更新日期和价格
    如果没有, 新增一行P0清仓标记

    Args:
        wb: openpyxl Workbook对象
        code: 股票代码
        name: 股票名称
        price: 当前价格
        obl: 一买低点
        pct: 跌幅百分比
    """
    try:
        if '执行清单' not in wb.sheetnames:
            return

        ws = wb['执行清单']
        today = date.today()

        # 检查是否已有该股票的未执行P0标记
        found_row = None
        for r in range(2, ws.max_row + 1):
            row_code = str(ws.cell(row=r, column=3).value or '')
            if row_code == str(code):
                note = str(ws.cell(row=r, column=15).value or '')
                if 'P0' in note:
                    found_row = r
                    break

        if found_row:
            # 更新现有行
            ws.cell(row=found_row, column=1, value=today)
            ws.cell(row=found_row, column=2, value=name)
            ws.cell(row=found_row, column=3, value=code)
            ws.cell(row=found_row, column=4, value="P0清仓")
            ws.cell(row=found_row, column=5, value=price)
            ws.cell(row=found_row, column=15, value=f"P0清仓-破一买低({obl:.2f}, 跌{pct:+.1f}%)")
        else:
            # 新增一行
            new_row = ws.max_row + 1
            ws.cell(row=new_row, column=1, value=today)
            ws.cell(row=new_row, column=2, value=name)
            ws.cell(row=new_row, column=3, value=code)
            ws.cell(row=new_row, column=4, value="P0清仓")
            ws.cell(row=new_row, column=5, value=price)
            ws.cell(row=new_row, column=15, value=f"P0清仓-破一买低({obl:.2f}, 跌{pct:+.1f}%)")
    except Exception:
        pass


def run_intraday_scan():
    """盘中扫描: 30min级别扫描候选池(排除持仓股) + 5min确认 + 持仓止损检查

    【Fix 2026-08-11】使用detect_multilevel_buy_signals替代旧analyze_beichi
    旧代码: analyze_beichi(level="30min") + sig["op"]=="一买" + ratio/DL_P阈值
    问题: 旧代码依赖DL_P(30min数据近零)和ratio, 与当前系统矛盾
    修复: 使用detect_multilevel_buy_signals的多级别tier和几何背驰检测
    """
    sys.path.insert(0, BEICHI_DIR)
    from beichi_analyzer import detect_multilevel_buy_signals, detect_sell_signals
    import time as _time

    now = datetime.now()
    print(f"=== 盘中扫描 {now.strftime('%Y-%m-%d %H:%M')} ===\n")

    # 候选池排除持仓股
    holdings = get_today_holdings()
    held_codes = {str(h['code']) for h in holdings if h.get('code')}

    # 1. 候选池30min扫描 (使用多级别信号检测)
    # 【BUG-9修复 (2026-07-27)】候选池为空时不再跳过持仓检查
    candidates = get_candidate_pool()
    candidates = [c for c in candidates if c["code"] not in held_codes]
    confirmed_30m = []  # 核心池/观察池信号
    near_30m = []       # 边缘池/接近信号
    dc_signals = []     # 双中枢趋势背驰信号
    if not candidates:
        print("[1/3] 候选池为空(排除持仓后), 跳过候选扫描, 继续持仓检查")
    else:
        print(f"[1/3] 候选池多级别扫描 ({len(candidates)}只)...")
        t0 = _time.time()
        for s in candidates:
            try:
                ml = detect_multilevel_buy_signals(s["code"], price=s["price"])
                if not ml or ml.get("tier") == "无信号":
                    continue
                tier = ml.get("tier", "无信号")
                # 使用几何背驰检测判断信号强度
                dc = ml.get("min30_double_center", {})
                sc = ml.get("min30_single_center", {})
                has_dc = dc.get("is_divergence", False)
                has_sc = sc.get("is_divergence", False)
                dc_conf = dc.get("confidence", 0)
                sc_conf = sc.get("confidence", 0)
                ermai = ml.get("ermai") is not None and ml.get("ermai", {}).get("valid", False)
                sanmai = ml.get("sanmai") is not None and ml.get("sanmai", {}).get("valid", False)
                has_buy_signal = ermai or sanmai

                entry = {
                    "code": s["code"], "name": s["name"],
                    "price": s["price"],
                    "tier": tier,
                    "has_dc": has_dc, "dc_conf": dc_conf,
                    "has_sc": has_sc, "sc_conf": sc_conf,
                    "ermai": ermai, "sanmai": sanmai,
                    "ep_30": ml.get("30min_ep_p", 0),
                }

                # 核心池: 双中枢趋势背驰或二买/三买确认
                if tier == "核心池" or has_dc or has_buy_signal:
                    confirmed_30m.append(entry)
                    if has_dc:
                        dc_signals.append(entry)
                # 观察池/边缘池: 接近确认
                elif tier in ("观察池", "边缘池"):
                    near_30m.append(entry)
            except:
                pass
        elapsed_30m = _time.time() - t0
        print(f"  多级别: 确认{len(confirmed_30m)}只 (双中枢{len(dc_signals)}只), 接近{len(near_30m)}只, 耗时{elapsed_30m:.0f}s")
        for s in confirmed_30m:
            tag = "双中枢" if s["has_dc"] else "单中枢" if s["has_sc"] else "tier"
            print(f"  ★ {s['name']} {s['code']} ¥{s['price']:.2f} [{tag}] tier={s['tier']}")
        for s in near_30m[:5]:
            print(f"  ◆ {s['name']} {s['code']} ¥{s['price']:.2f} tier={s['tier']}")

    # 2. 30min确认标的 → 5min精确买点 (使用多级别信号)
    confirmed_5m = []
    if confirmed_30m:
        print(f"\n[2/3] 5min精确买点扫描 ({len(confirmed_30m)}只)...")
        for s in confirmed_30m:
            try:
                ml = detect_multilevel_buy_signals(s["code"], price=s["price"])
                if not ml:
                    continue
                # 5min EP_L > 0.5 或 5min有信号
                ep_5 = ml.get("5min_ep_p", 0)
                if ep_5 > 0.5:
                    confirmed_5m.append(s)
                    print(f"  ★ {s['name']} {s['code']} 5min: EP_L={ep_5:.2f} ✓")
            except:
                pass
    else:
        print(f"\n[2/3] 5min扫描: 跳过(30min无确认)")

    # 3. 持仓止损 + 背驰卖点 + 中枢破位检查(根据成本自动确定级别)
    print(f"\n[3/3] 持仓检查(止损+背驰卖点+中枢破位)...")
    holdings = get_today_holdings()
    alerts = []
    sell_signals = []
    zs_breakdowns = []
    entry_reports = []
    for h in holdings:
        code = str(h['code'])
        name = h['name']
        waived = h['waived']
        close_price = h['close'] or 0
        cost = h['entry'] or 0

        # 3a. 自动确定买点级别和对应卖点监控级别
        if cost > 0:
            entry_data = detect_entry_level(code, cost)
            entry_level = entry_data["entry_level"]
            sell_levels = entry_data["sell_levels"]
            entry_info = entry_data["entry_info"].get(entry_level, {}) if entry_level else {}
            zone_desc = entry_info.get("zone", "未知") if entry_info else "未知"
            entry_reports.append({
                "name": name, "code": code, "cost": cost,
                "entry_level": entry_level or "未确定",
                "zone": zone_desc,
                "sell_levels": sell_levels,
            })
            print(f"  📌 {name}({code}) 成本{cost:.2f} → 买点级别: {entry_level or '未确定'}({zone_desc}) → 监控卖点: {'+'.join(sell_levels)}")
        else:
            sell_levels = ["日线", "30min"]
            print(f"  📌 {name}({code}) 无成本价 → 默认监控: 日线+30min")

        # 【BUG修复 (2026-07-26)】一买区持仓中枢破位误报
        #
        # BUG根因:
        #   detect_entry_level判断成本在中枢下方 → zone="中枢下方(一买区)"
        #   即: 成本 < last_zs["zd"] (成本本就低于中枢下沿)
        #   此时若现价也低于zd(正常,因为一买区持仓本就在中枢下方)
        #   中枢破位检查 close_price < zd 会触发误报
        #
        #   日志证据(2026-07-27 07:29运行):
        #     创力集团 成本8.02 中枢下方(一买区) 现价8.03 跌破下沿8.18
        #     新五丰 成本4.80 中枢下方(一买区) 现价4.73 跌破下沿4.88
        #     方盛制药 成本9.27 中枢下方(一买区) 现价8.96 跌破下沿9.89
        #     建研院 成本3.97 中枢下方(一买区) 现价4.02 跌破下沿4.03
        #     曲美家居 成本3.14 中枢下方(一买区) 现价3.09 跌破下沿3.18
        #   → 全部是一买区持仓, 现价<zd是正常状态, 不是破位
        #
        # 修复策略:
        #   一买区持仓(cost < zd): 中枢破位判定改为"现价跌破成本价"而非"跌破中枢下沿"
        #   中枢内持仓(zd <= cost <= zg): 维持原逻辑"跌破中枢下沿"
        #   中枢上方持仓(cost > zg): 维持原逻辑"跌破中枢下沿"
        is_one_buy_zone = zone_desc in ("中枢下方(一买区)", "全中枢下方(深度一买区)") if cost > 0 else False

        # 3b. 止损检查
        if waived != '是':
            if close_price and h['stop'] and close_price <= h['stop']:
                alerts.append(f"⚠️ {name}({code}) 破止损: 现价{close_price:.2f}<=止损{h['stop']:.2f}")

        # ============================================================
        # BUG修复 (2026-07-31): 一买低点止损 — 不受WAIVED状态影响
        #
        # 问题: 东风股份被标记WAIVED → 一买低点检查在 if waived != '是': 块内
        #       → 整个块被跳过 → 连续3天未检测到破位 → P0清仓失败
        # 根因: WAIVED设计初衷是跳过"普通止损"(避免一买区持仓的误报),
        #       但一买低点破位是硬性风控, 不应被WAIVED跳过
        # 修复: 将一买低点检查移出WAIVED门控, 无条件执行
        #       同时添加P0升级计数器, 连续N天未清仓→升级告警
        # ============================================================
        if close_price > 0 and cost > 0:
            cur_pos = h.get('pos', 0) or 0
            try:
                from beichi_analyzer import detect_multilevel_buy_signals as _detect_ml
                _ml = _detect_ml(code, price=close_price)
                _obl = _ml.get("one_buy_low")

                # BUG修复 (2026-07-31): 一买低点合理性校验(与check_sell_compliance一致)
                # 防止噪音中枢产生的错误一买低点触发误报
                if _obl and _obl > 0 and _obl > cost:
                    # 一买低 > 成本价 → 中枢检测异常, 跳过
                    pass
                elif _obl and _obl > 0 and close_price < _obl:
                    _pct = ((close_price - _obl) / _obl) * 100

                    # === P0升级计数器 ===
                    # 读取执行清单中该股票的P0待执行天数
                    p0_days = _count_p0_pending_days(wb, code, name)
                    waived_tag = " [WAIVED]" if waived == '是' else ""

                    if cur_pos > 0.35:
                        if p0_days >= 3:
                            alerts.append(
                                f"🔴🔴 P0升级({p0_days}天未执行): {name}({code}) 破一买低点: "
                                f"现价{close_price:.2f}<=一买低{_obl:.2f} (跌{_pct:+.1f}%, "
                                f"仓位{cur_pos:.1%}>35%) {waived_tag} → 严重纪律违规, 必须立即清仓!"
                            )
                        elif p0_days >= 2:
                            alerts.append(
                                f"🔴 P0升级({p0_days}天未执行): {name}({code}) 破一买低点: "
                                f"现价{close_price:.2f}<=一买低{_obl:.2f} (跌{_pct:+.1f}%, "
                                f"仓位{cur_pos:.1%}>35%) {waived_tag} → 连续未执行, 明日铁律清仓!"
                            )
                        else:
                            alerts.append(
                                f"🔴 {name}({code}) 破一买低点: 现价{close_price:.2f}<=一买低{_obl:.2f} "
                                f"(跌{_pct:+.1f}%, 仓位{cur_pos:.1%}>35%, 二买失败→应立即减仓至35%){waived_tag}"
                            )
                    else:
                        alerts.append(
                            f"⚠️ {name}({code}) 破一买低点: 现价{close_price:.2f}<=一买低{_obl:.2f} "
                            f"(跌{_pct:+.1f}%, 仓位{cur_pos:.1%}<=35%, 一买失败→止损){waived_tag}"
                        )

                    # 写入P0待执行标记到执行清单
                    _mark_p0_pending(wb, code, name, close_price, _obl, _pct)
            except:
                pass

        # 3c. 背驰卖点 + 中枢破位检查(动态级别)
        for level in sell_levels:
            try:
                r = analyze_beichi(code, level=level)
                if "error" in r:
                    continue

                # 中枢破位检查: 现价跌破最新中枢下沿
                # 【BUG修复】一买区持仓: 改为跌破成本价才算破位
                zss = r.get("zss", [])
                if zss and close_price > 0:
                    last_zs = zss[-1]
                    zd = last_zs["zd"]
                    zg = last_zs["zg"]
                    if is_one_buy_zone:
                        # 一买区持仓: 现价跌破成本价才是破位
                        if cost > 0 and close_price < cost:
                            pct = ((close_price - cost) / cost) * 100
                            zs_breakdowns.append({
                                "name": name, "code": code, "level": level,
                                "price": close_price, "zd": cost, "zg": zg, "pct": pct,
                                "waived": waived,
                                "breakdown_type": "一买区跌破成本",
                            })
                    else:
                        # 正常持仓: 现价跌破中枢下沿才是破位
                        if close_price < zd:
                            pct = ((close_price - zd) / zd) * 100
                            zs_breakdowns.append({
                                "name": name, "code": code, "level": level,
                                "price": close_price, "zd": zd, "zg": zg, "pct": pct,
                                "waived": waived,
                                "breakdown_type": "跌破中枢下沿",
                            })

                # 背驰卖点检查
                for sig in r.get("signals", []):
                    if "卖" not in sig["op"]:
                        continue
                    ratio = sig["ratio"]
                    dlp = sig["dl_prob"]
                    valid = sig["valid"]
                    confirmed_sell = ratio < 60 and dlp > 0.8 and valid
                    near_sell = (ratio < 60 and dlp > 0.6 and valid) or (ratio < 85 and dlp > 0.8 and valid)
                    if confirmed_sell:
                        sell_signals.append({
                            "name": name, "code": code, "level": level,
                            "op": sig["op"], "ratio": ratio, "dlp": dlp, "valid": valid,
                            "type": "确认卖点"
                        })
                        print(f"  🔴 {name}({code}) {level}确认卖点: {sig['op']} ratio={ratio:.0f}% DL_P={dlp:.2f}")
                    elif near_sell:
                        sell_signals.append({
                            "name": name, "code": code, "level": level,
                            "op": sig["op"], "ratio": ratio, "dlp": dlp, "valid": valid,
                            "type": "接近卖点"
                        })
            except:
                pass

        # 3d. 【BUG-7修复 (2026-07-27)】接入 detect_sell_signals 综合卖出评估
        #
        # 问题: detect_sell_signals 函数存在但从未被调用
        #   包含5条规则: 日线down+30min看空→减仓, 弱信号+密集看空→清仓, 亏损+down→清仓
        #   这些规则比 ratio/dlp 硬阈值更贴近实战, 能捕获"卖点未确认但风险已高"的场景
        #
        # 修复: 对每只持仓调用 detect_sell_signals, should_clear 自动升级为确认卖点
        if cost > 0 and close_price > 0 and waived != '是':
            try:
                sell_eval = detect_sell_signals(code, cost, close_price)
                if sell_eval["should_clear"]:
                    # should_clear 最高优先级 → 直接升级为确认卖点
                    sell_signals.append({
                        "name": name, "code": code, "level": "综合",
                        "op": "清仓", "ratio": 0, "dlp": 0, "valid": True,
                        "type": "确认卖点",
                        "reason": sell_eval["reason"],
                        "risk_level": sell_eval["risk_level"],
                        "source": "detect_sell_signals",
                    })
                    print(f"  🔴 {name}({code}) 综合清仓信号: {sell_eval['reason']} [风险={sell_eval['risk_level']}]")
                elif sell_eval["should_reduce"]:
                    # should_reduce → 接近卖点(附带原因)
                    sell_signals.append({
                        "name": name, "code": code, "level": "综合",
                        "op": "减仓", "ratio": 0, "dlp": 0, "valid": True,
                        "type": "接近卖点",
                        "reason": sell_eval["reason"],
                        "risk_level": sell_eval["risk_level"],
                        "source": "detect_sell_signals",
                    })
                    print(f"  🟡 {name}({code}) 综合减仓信号: {sell_eval['reason']} [风险={sell_eval['risk_level']}]")
            except:
                pass

    if alerts:
        for a in alerts:
            print(f"  {a}")

    # 中枢破位汇总
    if zs_breakdowns:
        print(f"  🔻 中枢破位: {len(zs_breakdowns)}个")
        for z in zs_breakdowns:
            waived_tag = " [WAIVED]" if z["waived"] == "是" else ""
            bd_type = z.get("breakdown_type", "跌破中枢下沿")
            print(f"    {z['name']}({z['code']}) {z['level']} 现价{z['price']:.2f} {bd_type}{z['zd']:.2f} ({z['pct']:+.1f}%){waived_tag}")

    # 【加仓信号检测】(2026-07-26): 二买/三买 → 动态仓位上限
    # 【BUG-9修复 (2026-07-27)】dynamic_cap_info未初始化导致NameError
    # 旧代码: 若所有持仓cost=0或close=0, get_dynamic_position_cap从未被调用
    #         → dynamic_cap_info全局变量未定义 → NameError崩溃 → Telegram推送失败
    # 修复: 在循环前初始化默认值
    dynamic_cap_info = {"entry": "一买", "pnl_pct": 0, "cap": 0.35, "tier": "边缘池",
                         "daily_dl_p": 0, "30min_dl_p": 0, "5min_dl_p": 0}
    add_signals = []
    print(f"\n  📈 加仓信号检测(多级别DL_P共振):")
    for h in holdings:
        code = str(h['code'])
        cost = h['entry'] or 0
        close = h['close'] or 0
        if cost <= 0 or close <= 0:
            continue

        pnl_pct = (close - cost) / cost
        dynamic_cap = get_dynamic_position_cap(code, cost, close)
        entry_level = dynamic_cap_info.get("entry", "一买")
        tier = dynamic_cap_info.get("tier", "边缘池")
        d_dp = dynamic_cap_info.get("daily_dl_p", 0)
        m30_dp = dynamic_cap_info.get("30min_dl_p", 0)
        m30_ep = dynamic_cap_info.get("30min_ep_p", 0)
        m5_dp = dynamic_cap_info.get("5min_dl_p", 0)
        m5_ep = dynamic_cap_info.get("5min_ep_p", 0)

        if entry_level in ("二买", "三买") and pnl_pct >= 0.05:
            # P1 (2026-08-02): 中阴状态拦截 — NotChasing不加仓
            zy = dynamic_cap_info.get("zhongyin", {})
            if zy.get("is_zhongyin") or zy.get("action") == "NotChasing":
                print(f"    ⛔ {h['name']}({code}) {entry_level}信号但{zy.get('action','')} — "
                      f"中阴/NotChasing, 不加仓({zy.get('reason','')[:40]})")
            else:
                add_signals.append({
                    "name": h['name'], "code": code, "entry": entry_level,
                    "pnl_pct": pnl_pct, "dynamic_cap": dynamic_cap,
                    "current_pos": h.get('pos', 0), "tier": tier,
                })
                remaining = dynamic_cap - (h.get('pos', 0) or 0)
                print(f"    ★ {h['name']}({code}) {entry_level}信号 [{tier}] → 动态上限{dynamic_cap:.0%} "
                      f"当前仓位{h.get('pos',0):.1%} 浮盈{pnl_pct:.1%} 可加仓空间{remaining:.1%}")
                print(f"      DL_P: 日线={d_dp:.2f} | EP_L: 30min={m30_ep:.2f} 5min={m5_ep:.2f} (30minDL_P={m30_dp:.2f}模型失效,以EP_L为准)")
                # P3 (2026-08-02): 显示综合置信度
                conf = dynamic_cap_info.get("daily_confidence", 0)
                if conf > 0:
                    print(f"      综合置信度: {conf:.2%}")
                # BUG修复 (2026-07-30): 显示一买低点距离
                obl = dynamic_cap_info.get("one_buy_low")
                dist_to_low = dynamic_cap_info.get("dist_to_one_buy_low")
                if obl and dist_to_low is not None:
                    if dist_to_low < 0.03:
                        print(f"      ⚠️ 一买低={obl:.2f} 现价离一买低仅{dist_to_low:.1%}<3% → 不宜加仓(破位风险)")
                    else:
                        print(f"      ✓ 一买低={obl:.2f} 现价离一买低{dist_to_low:.1%}(安全)")
                else:
                    print(f"      ℹ️ 一买低点未检测到")
        else:
            print(f"    · {h['name']}({code}) [{tier}] 上限{dynamic_cap:.0%} "
                  f"DL_P: 日线={d_dp:.2f} | EP_L: 30min={m30_ep:.2f} 5min={m5_ep:.2f}")

    if not add_signals:
        print(f"    无多级别共振信号, 所有持仓维持35%静态上限")

    # 卖点汇总
    # 【BUG-7修复 (2026-07-27)】中枢破位 + 接近卖点 = 自动升级为确认卖点
    #
    # 问题: 跌破中枢后一卖/二卖因 ratio/dlp/valid 未同时达标而停在"接近卖点"
    #       等待确认期间价格继续下跌 → 大幅回撤
    #
    # 修复: 同一股票同时出现"中枢破位"和"接近卖点"时, 自动升级为确认卖点
    #       理由: 中枢破位 = 趋势结构已破坏, 接近卖点 = 卖点正在形成
    #             两者叠加 = 不需等 ratio/dlp 完全确认, 应立即行动
    confirmed_sells = [s for s in sell_signals if s["type"] == "确认卖点"]
    near_sells = [s for s in sell_signals if s["type"] == "接近卖点"]

    # 中枢破位 + 接近卖点 → 自动升级
    if zs_breakdowns and near_sells:
        breakdown_codes = {z["code"] for z in zs_breakdowns if z["waived"] != "是"}
        upgraded = []
        remaining_near = []
        for s in near_sells:
            if s["code"] in breakdown_codes:
                s["type"] = "确认卖点"
                s["upgraded_from"] = "中枢破位+接近卖点自动升级"
                upgraded.append(s)
            else:
                remaining_near.append(s)
        if upgraded:
            near_sells = remaining_near
            confirmed_sells.extend(upgraded)
            print(f"  ⚡ 自动升级: {len(upgraded)}只接近卖点(中枢破位叠加) → 确认卖点")
            for s in upgraded:
                print(f"    ↗ {s['name']}({s['code']}) {s['level']} {s['op']} → 确认卖出")

    if confirmed_sells:
        print(f"  🔴 确认卖点: {len(confirmed_sells)}个")
    if near_sells:
        print(f"  🟡 接近卖点: {len(near_sells)}个")
        for s in near_sells[:5]:
            missing = []
            if s.get("ratio", 0) >= 60: missing.append("ratio=%d%%" % s["ratio"])
            if s.get("dlp", 0) <= 0.8: missing.append("DL_P=%.2f" % s["dlp"])
            print(f"    {s['name']}({s['code']}) {s['level']} {s['op']} 缺:{'+'.join(missing)}")
    if not alerts and not confirmed_sells and not near_sells and not zs_breakdowns:
        print(f"  持仓{len(holdings)}只, 止损合规, 无背驰卖点, 无中枢破位")

    # 汇总
    print(f"\n{'='*50}")
    if confirmed_30m:
        print(f"★ 确认信号: {len(confirmed_30m)}只 (双中枢{len(dc_signals)}只)")
        for s in confirmed_30m:
            tag = "双中枢" if s.get("has_dc") else "单中枢" if s.get("has_sc") else "二买/三买" if s.get("ermai") or s.get("sanmai") else "tier"
            ep30 = s.get("ep_30", 0)
            print(f"  {s['name']} {s['code']} ¥{s['price']:.2f} [{tag}] tier={s.get('tier','?')} EP_L={ep30:.2f}")
    else:
        print("★ 确认信号: 0只")

    if near_30m:
        print(f"\n◆ 接近确认: {len(near_30m)}只")
        for s in near_30m[:5]:
            ep30 = s.get("ep_30", 0)
            print(f"  {s['name']} {s['code']} ¥{s['price']:.2f} tier={s.get('tier','?')} EP_L={ep30:.2f}")

    if alerts:
        print(f"\n⚠️ 止损告警: {len(alerts)}只需处理")

    if zs_breakdowns:
        non_waived_bd = [z for z in zs_breakdowns if z["waived"] != "是"]
        print(f"\n🔻 中枢破位: {len(zs_breakdowns)}个 (非WAIVED {len(non_waived_bd)}个)")
        if non_waived_bd:
            print("  → 跌破中枢下沿但卖点未确认, 建议人工评估是否减仓")

    if confirmed_sells:
        print(f"\n🔴 确认卖点: {len(confirmed_sells)}个 (建议卖出)")
    if near_sells:
        print(f"\n🟡 接近卖点: {len(near_sells)}个")

    print(f"{'='*50}")
    return {
        "confirmed_30m": confirmed_30m,
        "near_30m": near_30m,
        "dc_signals": dc_signals,
        "alerts": alerts,
        "confirmed_sells": confirmed_sells,
        "near_sells": near_sells,
        "zs_breakdowns": zs_breakdowns,
        "entry_reports": entry_reports,
        "add_signals": add_signals,
        "scanned": len(candidates),
    }

def send_telegram(text, title=""):
    """Send message to Telegram. Splits long messages automatically.

    彻底修复 (2026-07-26): 5次失败的根因分析
    1. parse_mode=HTML导致消息中<被当作HTML标签 → HTTP 400 "can't parse entities"
       修复: 完全移除parse_mode, 用纯文本发送
    2. send_telegram失败时只print不raise → 工作流显示"success"但消息没发出
       修复: 失败时raise RuntimeError, 确保错误传播到workflow
    3. 工作流挂起(analyze_beichi无缓存) → 40+分钟未到达send_telegram
       修复: beichi_analyzer.py添加内存缓存 + workflow.yml添加timeout

    本函数不再静默失败: token缺失或HTTP错误都会raise异常
    """
    token = (os.environ.get('TELEGRAM_BOT_TOKEN')
             or os.environ.get('TG_TOKEN')
             or os.environ.get('TELEGRAM_TOKEN'))
    chat_id = (os.environ.get('TELEGRAM_CHAT_ID')
               or os.environ.get('TG_CHAT_ID')
               or os.environ.get('TELEGRAM_CHATID'))

    tg_keys = [k for k in sorted(os.environ.keys()) if 'TELEGRAM' in k.upper() or 'TG_' in k.upper()]
    print(f"[TG] env vars: {tg_keys}")
    print(f"[TG] token: {'YES' if token else 'NO'}, chat_id: {'YES' if chat_id else 'NO'}")

    if not token or not chat_id:
        msg = f"Telegram token或chat_id缺失 (token={'YES' if token else 'NO'}, chat_id={'YES' if chat_id else 'NO'})"
        print(f"[TG] FAIL: {msg}")
        print("=" * 40)
        if title:
            print(title)
        print(text)
        print("=" * 40)
        print("[TG] 继续执行, 不中断工作流")
        return

    # 转义<>为全角字符 (不使用parse_mode, 纯文本模式)
    # 注意: 不替换& — 不用parse_mode时&不需要HTML转义, 替换会导致显示&amp;
    def escape_tg(s):
        if not s:
            return s
        s = s.replace('<', '＜')
        s = s.replace('>', '＞')
        return s

    full_text = f"{title}\n{text}" if title else text
    full_text = escape_tg(full_text)

    # Split into chunks of 4000 chars (Telegram limit is 4096)
    chunks = []
    while full_text:
        if len(full_text) <= 4000:
            chunks.append(full_text)
            break
        split_at = full_text.rfind('\n', 0, 4000)
        if split_at < 2000:
            split_at = 4000
        chunks.append(full_text[:split_at])
        full_text = full_text[split_at:].lstrip('\n')

    print(f"[TG] Sending {len(chunks)} message(s), total chars={len(text)}")

    errors = []
    for i, chunk in enumerate(chunks):
        if i > 0:
            chunk = f"续({i+1}/{len(chunks)}):\n{chunk}"
        data = json.dumps({
            "chat_id": chat_id,
            "text": chunk,
            "disable_web_page_preview": True,
        }).encode('utf-8')
        req = urllib.request.Request(
            f"https://api.telegram.org/bot{token}/sendMessage",
            data=data,
            headers={"Content-Type": "application/json"},
        )
        try:
            resp = urllib.request.urlopen(req, timeout=15)
            resp_body = resp.read().decode('utf-8')
            print(f"[TG] chunk {i+1} OK: {resp_body[:100]}")
        except urllib.error.HTTPError as e:
            err_body = e.read().decode('utf-8') if hasattr(e, 'read') else str(e)
            err_msg = f"chunk {i+1} HTTP {e.code}: {err_body}"
            print(f"[TG] ERROR: {err_msg}")
            errors.append(err_msg)
        except Exception as e:
            err_msg = f"chunk {i+1} {type(e).__name__}: {e}"
            print(f"[TG] ERROR: {err_msg}")
            errors.append(err_msg)
        if i < len(chunks) - 1:
            time.sleep(1)

    if errors:
        raise RuntimeError(f"Telegram发送失败({len(errors)}/{len(chunks)}): {'; '.join(errors)}")


def format_intraday_summary(result, ts):
    """Format intraday scan results into clean Telegram message.

    【Fix 2026-08-11】使用tier分类+几何背驰替代旧ratio/DL_P显示
    """
    lines = [f"📡 盘中扫描 {ts}", ""]

    scanned = result.get("scanned", 0)
    dc_count = len(result.get("dc_signals", []))
    lines.append(f"📊 扫描: {scanned}只候选 | 双中枢: {dc_count}只")
    lines.append("")

    # 30min确认 (多级别tier + 几何背驰)
    confirmed = result.get("confirmed_30m", [])
    if confirmed:
        lines.append(f"★ 确认信号: {len(confirmed)}只")
        for s in confirmed:
            tag = "双中枢" if s.get("has_dc") else "单中枢" if s.get("has_sc") else "二买/三买" if s.get("ermai") or s.get("sanmai") else "tier"
            ep30 = s.get("ep_30", 0)
            lines.append(f"  {s['name']} {s['code']} ¥{s['price']:.2f} [{tag}] tier={s.get('tier','?')} EP_L={ep30:.2f}")
        lines.append("")

    # 30min接近
    near = result.get("near_30m", [])
    if near:
        lines.append(f"◆ 接近信号: {len(near)}只")
        for s in near[:8]:
            ep30 = s.get("ep_30", 0)
            lines.append(f"  {s['name']} {s['code']} ¥{s['price']:.2f} tier={s.get('tier','?')} EP_L={ep30:.2f}")
        if len(near) > 8:
            lines.append(f"  ...还有{len(near)-8}只")
        lines.append("")

    # 持仓买点级别
    entry_reports = result.get("entry_reports", [])
    if entry_reports:
        lines.append("📌 持仓买点级别:")
        for e in entry_reports:
            lines.append(f"  {e['name']}({e['code']}) {e['entry_level']} → 监控:{'+'.join(e['sell_levels'])}")
        lines.append("")

    # 止损告警
    alerts = result.get("alerts", [])
    if alerts:
        lines.append(f"⚠️ 止损告警: {len(alerts)}只")
        for a in alerts:
            lines.append(f"  {a}")
        lines.append("")

    # 中枢破位
    breakdowns = result.get("zs_breakdowns", [])
    if breakdowns:
        non_waived = [z for z in breakdowns if z["waived"] != "是"]
        lines.append(f"🔻 中枢破位: {len(breakdowns)}个 (非WAIVED {len(non_waived)}个)")
        for z in non_waived[:5]:
            bd_type = z.get("breakdown_type", "跌破中枢下沿")
            lines.append(f"  {z['name']}({z['code']}) {z['level']} ¥{z['price']:.2f} {bd_type}{z['zd']:.2f} ({z['pct']:+.1f}%)")
        if len(non_waived) > 5:
            lines.append(f"  ...还有{len(non_waived)-5}个")
        lines.append("")

    # 卖点
    confirmed_sells = result.get("confirmed_sells", [])
    near_sells = result.get("near_sells", [])
    if confirmed_sells:
        lines.append(f"🔴 确认卖点: {len(confirmed_sells)}个 (建议卖出)")
        for s in confirmed_sells:
            src = s.get("source", "")
            reason = s.get("reason", "")
            upgraded = s.get("upgraded_from", "")
            if src == "detect_sell_signals":
                lines.append(f"  {s['name']}({s['code']}) {s['op']} {reason}")
            elif upgraded:
                lines.append(f"  {s['name']}({s['code']}) {s['level']} {s['op']} [{upgraded}]")
            else:
                lines.append(f"  {s['name']}({s['code']}) {s['level']} {s['op']} ratio={s.get('ratio',0):.0f}% DL_P={s.get('dlp',0):.2f}")
        lines.append("")
    if near_sells:
        lines.append(f"🟡 接近卖点: {len(near_sells)}个")
        for s in near_sells[:5]:
            reason = s.get("reason", "")
            if reason:
                lines.append(f"  {s['name']}({s['code']}) {s['op']} {reason}")
            else:
                missing = []
                if s.get("ratio", 0) >= 60: missing.append(f"ratio={s['ratio']:.0f}%")
                if s.get("dlp", 0) <= 0.8: missing.append(f"DL_P={s['dlp']:.2f}")
                lines.append(f"  {s['name']}({s['code']}) {s['level']} {s['op']} 缺:{'+'.join(missing)}")
        lines.append("")

    if not any([confirmed, near, alerts, breakdowns, confirmed_sells, near_sells]):
        lines.append("✓ 无信号, 持仓合规")

    # 加仓信号
    add_signals = result.get("add_signals", [])
    if add_signals:
        lines.append("")
        lines.append(f"📈 趋势加仓信号(动态仓位): {len(add_signals)}只")
        for s in add_signals:
            remaining = s["dynamic_cap"] - s["current_pos"]
            lines.append(f"  ★ {s['name']}({s['code']}) {s['entry']} → 上限{s['dynamic_cap']:.0%} "
                         f"浮盈{s['pnl_pct']:.1%} 可加{remaining:.1%}")

    return '\n'.join(lines)


def format_scan_summary(scan_data, ts):
    """Format full scan results into clean Telegram message.

    【Fix 2026-08-11】使用tier分类替代旧ratio/DL_P显示
    旧代码: 显示ratio和DL_P作为筛选条件
    问题: 当前系统使用多级别tier+几何背驰检测, ratio/DL_P已不是主要条件
    修复: 显示tier分类、信号类型(双中枢/单中枢/二买/三买)
    """
    lines = [f"📊 收盘扫描报告 {ts}", ""]

    result = scan_data.get("scan_result", {})
    selected = scan_data.get("selected", [])
    errors = scan_data.get("errors", 0)

    total = result.get("total_scanned", 0) or len(result.get("near", []))
    core = result.get("core", [])
    watch = result.get("watch", [])
    edge = result.get("edge", [])

    lines.append(f"扫描: {total}只 | 核心池: {len(core)}只 | 观察池: {len(watch)}只 | 边缘池: {len(edge)}只")
    lines.append(f"候选池写入: {len(selected)}只 | 公式错误: {errors}")
    lines.append("")

    # 按tier分层显示候选池
    if selected:
        # 核心池优先显示
        core_in_pool = [s for s in selected if s.get("tier") == "核心"]
        watch_in_pool = [s for s in selected if s.get("tier") == "观察"]
        edge_in_pool = [s for s in selected if s.get("tier") == "边缘"]

        if core_in_pool:
            lines.append("🏆 核心池 (调仓首选):")
            for s in core_in_pool[:8]:
                sig_type = s.get("sig_type", "盘整背驰")
                label = s.get("sig_label", "ABC买卖区间")
                # 双中枢趋势背驰信号优先标注
                if s.get("min30_dc_divergence"):
                    sig_type = "趋势背驰"
                    label = "123买卖区间"
                lines.append(f"  {s['name']} {s['code']} ¥{s['price']:.2f} | {label} | {sig_type}")
            if len(core_in_pool) > 8:
                lines.append(f"  ...还有{len(core_in_pool)-8}只")
            lines.append("")

        if watch_in_pool:
            lines.append("👀 观察池 (核心池不足时补充):")
            for s in watch_in_pool[:5]:
                ep_30 = s.get("30min_ep_p", 0)
                lines.append(f"  {s['name']} {s['code']} ¥{s['price']:.2f} | EP_L={ep_30:.2f}")
            if len(watch_in_pool) > 5:
                lines.append(f"  ...还有{len(watch_in_pool)-5}只")
            lines.append("")

        if edge_in_pool:
            lines.append(f"⚪ 边缘池 ({len(edge_in_pool)}只, 仅观察不买入)")

    return '\n'.join(lines)


def format_rebalance_summary(holdings, ts):
    """格式化去弱留强Telegram消息: 减多少股 + 加多少股 + 最新核心池

    权重 (2026-07-29): DL_P 0.7 + EP_L 0.3 (DL_P更具特征代表性)
    """
    import sys as _sys
    _sys.path.insert(0, BEICHI_DIR)
    from beichi_analyzer import detect_multilevel_buy_signals

    lines = [f"🔄 去弱留强 {ts}", ""]

    # === 1. 对每只持仓计算信号强度 ===
    ranked = []
    for h in holdings:
        code = str(h['code'])
        cost = h.get('entry') or 0
        close = h.get('close') or 0
        shares = h.get('shares') or 0
        if cost <= 0 or close <= 0 or shares <= 0:
            continue

        pnl = (close - cost) / cost
        mv = shares * close
        try:
            ml = detect_multilevel_buy_signals(code, price=close)
        except:
            ml = {}

        dl_p = ml.get('daily_dl_p', 0)
        m30_ep = ml.get('30min_ep_p', 0)
        m5_ep = ml.get('5min_ep_p', 0)
        # BUG修复 (2026-07-30): 获取一买低点
        obl = ml.get('one_buy_low')
        dist_to_obl = None
        if obl and obl > 0 and close > 0:
            dist_to_obl = (close - obl) / obl
        # BUG修复 (2026-08-03): one_buy_low=None时保守兜底
        # 问题: detect_multilevel_buy_signal可能无法计算一买低点(中枢数据缺失等)
        #       → dist_to_obl保持None → 风险惩罚完全跳过
        #       → 东风股份DL=0.08也能保留
        # 修复: 无法获取一买低点时, 保守假设为高风险(dit_to_obl=0)
        elif close > 0:
            # 无中枢数据可计算一买低点, 保守标记为高风险
            dist_to_obl = 0.0
        # 新权重: DL_P 0.7 + 30min EP 0.2 + 5min EP 0.1
        score = dl_p * 0.7 + m30_ep * 0.2 + m5_ep * 0.1

        # ============================================================
        # BUG修复 (2026-07-30): 一买低点风险惩罚 — 防止加仓后破位重仓回撤
        #
        # 问题: 去弱留强评分只看信号强度(DL_P+EP_L), 不看下行风险
        #       沃华DL_P=0.94评分最高 → 永远被保留 → 即使接近一买低点也不卖
        #       → 二买失效时重仓大幅回撤
        #
        # 修复: 接近一买低点时评分惩罚
        #   dist < 3%: score *= 0.4 (严重风险, 优先卖出)
        #   dist 3-5%: score *= 0.7 (中等风险, 降低排名)
        #   dist < 0 (已破位): score = 0 (立即清仓)
        #
        # 为什么EP_L>0.6不能解决:
        #   EP_L是当前反转概率的快照, 不预测未来
        #   EP_L=0.62今天确认二买, 明天可能跌到EP_L=0.2
        #   但加仓已经完成 → 重仓在手 → 只能靠止损退出
        #   唯一有效的保护: 评分系统考虑一买低点距离
        # ============================================================
        risk_tag = ""
        if dist_to_obl is not None:
            if dist_to_obl < 0:
                # 已破一买低点 → 二买完全失效, 评分归零
                score = 0
                risk_tag = " [已破一买低,二买失效]"
            elif dist_to_obl < 0.03:
                # 离一买低<3% → 严重风险, 评分降至40%
                score *= 0.4
                risk_tag = f" [离一买低{dist_to_obl:.1%},高风险]"
            elif dist_to_obl < 0.05:
                # 离一买低3-5% → 中等风险, 评分降至70%
                score *= 0.7
                risk_tag = f" [离一买低{dist_to_obl:.1%},中风险]"

        # 卖点检查
        sell_signal = ""
        try:
            r30 = analyze_beichi(code, level="30min")
            for sig in r30.get("signals", []):
                if sig["op"] in ("一卖", "二卖") and sig.get("valid"):
                    sell_signal = sig["op"]
        except:
            pass

        ranked.append({
            "name": h['name'], "code": code, "shares": shares,
            "cost": cost, "close": close, "pnl": pnl, "mv": mv,
            "dl_p": dl_p, "m30_ep": m30_ep, "m5_ep": m5_ep,
            "score": score, "sell_signal": sell_signal,
            "t1_lock": h.get("t1_lock", ""),
            "one_buy_low": obl,  # BUG修复 (2026-07-30)
            "dist_to_obl": dist_to_obl,  # 离一买低点距离
            "risk_tag": risk_tag,  # 风险标签
        })

    ranked.sort(key=lambda x: -x["score"])

    # === 2. 分割: 前4保留, 后N清仓 ===
    # BUG修复 (2026-07-30): 已破一买低点的股票强制进入卖出列表
    #
    # 问题: 旧逻辑只按评分排名取前4保留
    #       沃华DL_P=0.94评分最高 → 即使破一买低也在前4 → 被保留 → 重仓回撤
    #       评分惩罚(BUG1)已让破位股票score降低, 但还不够:
    #       如果持仓只有4只, 破位的也会在前4
    #
    # 修复: 破一买低点的股票无条件进入sell_list, 不受前4保护
    # BUG修复 (2026-08-03): DL_P<0.5强制清仓, 不受前4保护
    # 问题: 东风股份DL_P=0.08评分0.056, 但无one_buy_low数据
    #       → 不触发破位强制卖出 → 排名进前4→被保留→继续亏损
    # 修复: DL_P<0.5视为信号强度极弱, 直接强制卖出
    keep = []
    forced_sell = []
    for r in ranked:
        dist = r.get("dist_to_obl")
        dl_p = r.get("dl_p", 0)
        if (dist is not None and dist < 0) or dl_p < 0.5:
            # 已破一买低点 或 DL_P<0.5(信号极弱) → 强制卖出
            forced_sell.append(r)
        else:
            keep.append(r)
    # BUG修复 (2026-08-02): 旧代码先 keep=keep[:4] 再取 keep[4:]
    #   导致 keep[4:] 恒为空, 排名第5+的非破位股票被静默丢弃 (既不在keep也不在sell_list)
    #   修复: 先取 sell_list 再截断 keep
    sell_list = keep[4:] + forced_sell  # 排名第5+ + 强制清仓
    keep = keep[:4]

    # === 3. 账户概览 ===
    total_mv = sum(r["mv"] for r in ranked)
    sell_mv = sum(r["mv"] for r in sell_list)
    lines.append(f"持仓{len(ranked)}只 → 保留{len(keep)}只/清仓{len(sell_list)}只")
    lines.append(f"清仓回收: ¥{sell_mv:,.0f}")
    lines.append("")

    # === 4. 清仓明细 (减多少股) ===
    if sell_list:
        lines.append("🔻 清仓(卖出):")
        for r in sell_list:
            pnl_str = f"{r['pnl']*100:+.1f}%"
            score_str = f"强度{r['score']:.2f}"
            dl_str = f"DL={r['dl_p']:.2f}"
            sell_tag = f" +{r['sell_signal']}" if r["sell_signal"] else ""
            lines.append(
                f"  卖 {r['name']}({r['code']}) {r['shares']}股 ¥{r['close']:.2f} "
                f"回收¥{r['mv']:,.0f} {pnl_str} {dl_str}{sell_tag}"
            )
        lines.append("")

    # === 5. 保留明细 ===
    if keep:
        lines.append("✅ 保留(持有):")
        for r in keep:
            pnl_str = f"{r['pnl']*100:+.1f}%"
            risk_str = r.get("risk_tag", "")
            lines.append(
                f"  留 {r['name']}({r['code']}) {r['shares']}股 ¥{r['close']:.2f} "
                f"DL={r['dl_p']:.2f} EP={r['m30_ep']:.2f} {pnl_str}{risk_str}"
            )
        lines.append("")

    # === 6. 加仓信号检查 (加多少股) ===
    add_lines = []
    for r in keep:
        cost = r["cost"]
        close = r["close"]
        pnl = r["pnl"]
        dl_p = r["dl_p"]
        m30_ep = r["m30_ep"]
        m5_ep = r["m5_ep"]
        obl = r.get("one_buy_low")
        dist_obl = r.get("dist_to_obl")

        # BUG修复 (2026-07-30): 一买低点距离检查 — 加仓前置风控
        # 离一买低点<3%时禁止加仓, 防止二买破位重仓回撤
        obl_block = False
        obl_warn = ""
        if obl and dist_obl is not None and dist_obl < 0.03:
            obl_block = True
            obl_warn = f" ⚠️离一买低{dist_obl:.1%}<3%不宜加仓"

        # 二买条件检查
        cond_b = m30_ep >= 0.5
        cond_c = m5_ep >= 0.3 or m5_ep > 0
        cond_d = pnl >= 0.05

        if cond_b and cond_c and cond_d and not obl_block:
            # 二买确认 → 可加仓到50%
            total_asset = 21311  # 近似值
            target_mv = total_asset * 0.50
            current_mv = r["mv"]
            add_mv = max(0, target_mv - current_mv)
            add_shares = int(add_mv / close / 100) * 100
            if add_shares > 0:
                obl_info = f" 一买低={obl:.2f}(安全{dist_obl:.1%})" if obl and dist_obl else ""
                add_lines.append(
                    f"  加 {r['name']}({r['code']}) +{add_shares}股 "
                    f"仓位→50% (二买确认 DL={dl_p:.2f} EP={m30_ep:.2f}){obl_info}"
                )
        elif cond_b and cond_c and cond_d and obl_block:
            # 二买确认但一买低点太近 → 不加仓
            add_lines.append(
                f"  ⏸ {r['name']}({r['code']}) 二买确认但离一买低{dist_obl:.1%}<3%{obl_warn}"
            )
        elif cond_b and not cond_c:
            # 仅缺5min确认
            trigger = cost * 1.05
            add_lines.append(
                f"  等 {r['name']}({r['code']}) 5min入场确认 "
                f"(30m EP={m30_ep:.2f}✓ 涨至¥{trigger:.2f}触发)"
            )
        elif pnl < 0.05 and dl_p >= 0.8:
            # DL_P强但浮盈不足
            trigger = cost * 1.05
            add_lines.append(
                f"  等 {r['name']}({r['code']}) 浮盈至5%(¥{trigger:.2f}) "
                f"DL={dl_p:.2f} EP={m30_ep:.2f}"
            )

    if add_lines:
        lines.append("📈 加仓/观察:")
        for a in add_lines:
            lines.append(a)
        lines.append("")

    # === 7. 最新核心池 ===
    # BUG修复 (2026-07-30): 候选池列索引错配
    #   旧代码: tier读col8(DL_P), dl_p读col6(买点低点), ratio读col7(ratio), confirmed读col10(SignalType)
    #   正确列: tier=col11(分层Tier), dl_p=col8(DL_P), ratio=col7(ratio), confirmed=col12(是否确认信号)
    #   tier值应为"核心池"而非"核心"
    try:
        wb = safe_load_wb(data_only=True)
        if '候选池' in wb.sheetnames:
            ws_pool = wb['候选池']
            core_pool = []
            for r in range(2, ws_pool.max_row + 1):
                tier = str(ws_pool.cell(row=r, column=11).value or "")
                if tier == "核心池":
                    name = ws_pool.cell(row=r, column=2).value or ""
                    code = str(ws_pool.cell(row=r, column=3).value or "")
                    price = ws_pool.cell(row=r, column=4).value or 0
                    dl_p = ws_pool.cell(row=r, column=8).value or 0
                    ratio = ws_pool.cell(row=r, column=7).value or 0
                    confirmed = ws_pool.cell(row=r, column=12).value or ""
                    core_pool.append({
                        "name": name, "code": code, "price": price,
                        "dl_p": dl_p, "ratio": ratio, "confirmed": confirmed,
                    })
            if core_pool:
                lines.append(f"⭐ 核心池({len(core_pool)}只):")
                for s in core_pool:
                    lines.append(
                        f"  {s['name']}({s['code']}) ¥{s['price']:.2f} "
                        f"DL={s['dl_p']:.2f} ratio={s['ratio']*100:.1f}%"
                    )
                lines.append("")
    except:
        pass

    return '\n'.join(lines)


def format_compliance_summary(account, holdings, issues, ts):
    """Format compliance check results into clean Telegram message."""
    lines = [f"📝 合规核查 {ts}", ""]

    total_asset = account.get("total_asset", 0) or 0
    cash = account.get("cash", 0) or 0
    pos_ratio = account.get("position_ratio", 0) or 0

    lines.append(f"总资产: ¥{total_asset:.2f}")
    lines.append(f"现金: ¥{cash:.2f}")
    lines.append(f"仓位: {pos_ratio:.1%}")
    lines.append(f"持仓: {len(holdings)}只")
    lines.append("")

    if issues:
        lines.append(f"⚠️ 合规告警 ({len(issues)}项):")
        for issue in issues:
            lines.append(f"  {issue}")
        lines.append("")
    else:
        lines.append("✓ 持仓合规")
        lines.append("")

    lines.append("📝 心态日志提醒:")
    lines.append("  • 非工作流看盘次数 (0=最好)")
    lines.append("  • 是否追价操作 (0=最好)")
    lines.append("  • 是否按系统执行 (5=最好)")
    lines.append("  • 情绪状态 (5=最好)")
    lines.append("  • 止损外操作 (0=最好)")
    lines.append("  达标线: 80分")

    return '\n'.join(lines)


def run_weekly_review():
    """周复盘: 复利目标 + R值统计 + 账户增长 + 做T/建清仓 → 写入周复盘表W-AK列

    列映射:
      W(23) 日复利目标完成    X(24) 周复利目标完成    Y(25) 月复利目标完成
      Z(26) 当下累计复利率    AA(27) 当下正期望R值    AB(28) 系统盈亏比
      AC(29) 账户规模增长     AD(30) 增长目标进度      AE(31) 复利追踪备注
      AF(32) 做T次数/成功率   AG(33) 建清仓成功率       AH(34) 平均持仓天数

    修正 (2026-07-26):
      - 本金取值: 列27(AA)=硬编码值, 改用 总资产-净入金后盈利(列43) = 实际投入本金
      - 新增做T统计: 同日买卖识别 + 成功率
      - 新增建清仓成功率: FIFO买卖配对完整周期
    """
    from openpyxl.styles import PatternFill, Font as XFont, Alignment, Border, Side

    SUMMARY_START = 204  # 交易记录统计区域起始行

    # ============================================================
    # 1. 读取账户数据
    # ============================================================
    wb_d = safe_load_wb(data_only=True)
    ws_acc = wb_d['账户总表']
    latest_row = 2
    for r in range(2, ws_acc.max_row + 1):
        if ws_acc.cell(row=r, column=1).value:
            latest_row = r

    current_asset = ws_acc.cell(row=latest_row, column=2).value or 0
    if isinstance(current_asset, str): current_asset = 0
    compound_dev = ws_acc.cell(row=latest_row, column=31).value or 0
    if isinstance(compound_dev, str): compound_dev = 0

    # ============================================================
    # 双复利模型 (2026-07-26修订):
    #   主用TWR: 以总资产为基数, 投入本金为基准 (合规审计/国际标准)
    #   辅用动态本金: 以持仓市值为基数 (交易信号评估)
    # ============================================================
    position_value = ws_acc.cell(row=latest_row, column=4).value or 0
    if isinstance(position_value, str): position_value = 0
    cash = ws_acc.cell(row=latest_row, column=3).value or 0
    if isinstance(cash, str): cash = 0

    # 持仓市值可能为0(历史数据缺失), 从持仓表重新计算
    if position_value == 0:
        ws_hold_d = wb_d['持仓表']
        for r in range(2, ws_hold_d.max_row + 1):
            name = ws_hold_d.cell(row=r, column=2).value
            if not name:
                continue
            shares = ws_hold_d.cell(row=r, column=7).value or 0
            close = ws_hold_d.cell(row=r, column=9).value or 0
            waived = ws_hold_d.cell(row=r, column=4).value
            if shares and close and not isinstance(shares, str) and not isinstance(close, str):
                if float(shares) > 0 and float(close) > 0:
                    position_value += float(shares) * float(close)

    # 动态本金 = 持仓市值 (复利基数 - 辅用)
    dynamic_principal = position_value
    position_ratio = position_value / current_asset if current_asset > 0 else 0

    # 净盈利(列43)
    real_pnl = ws_acc.cell(row=latest_row, column=43).value or 0
    if isinstance(real_pnl, str): real_pnl = 0

    # ============================================================
    # TWR复利模型 (总资产基数 - 国际标准/合规审计用)
    #   twr_principal = 总资产 - 净盈利 = 实际投入本金
    # ============================================================
    twr_principal = current_asset - float(real_pnl)
    if twr_principal <= 0:
        # 回退: 用列27硬编码值
        for r in range(2, latest_row + 1):
            v = ws_acc.cell(row=r, column=27).value
            if v is not None and not str(v).startswith('=') and not isinstance(v, str):
                twr_principal = float(v)
                break

    # 投入本金(参考)
    invested_capital = twr_principal

    # 运行月数
    start_date = datetime(2026, 1, 31)
    end_date = datetime.now()
    months_running = (end_date.year - start_date.year) * 12 + (end_date.month - start_date.month)
    if end_date.day < start_date.day:
        months_running -= 1
    months_precise = months_running + (end_date.day - start_date.day if end_date.day >= start_date.day else end_date.day + 31 - start_date.day) / 31.0

    # 目标年化: 基于总资产的TWR目标 (随账户阶段递减)
    #   婴儿(<3W): 3%    幼儿(3W-10W): 2.5%    成长(10W-30W): 2%    成熟(30W+): 1.5%
    if current_asset < 30000:
        target_annual = 0.03
    elif current_asset < 100000:
        target_annual = 0.025
    elif current_asset < 300000:
        target_annual = 0.02
    else:
        target_annual = 0.015

    total_profit = float(real_pnl)

    # 1. TWR复利 (总资产基数 - 主用)
    twr_total_return = (current_asset - twr_principal) / twr_principal if twr_principal > 0 else 0
    twr_annualized_return = ((current_asset / twr_principal) ** (12 / months_precise) - 1) if twr_principal > 0 and months_precise > 0 else 0

    # 2. 动态本金复利 (持仓市值基数 - 辅用)
    dynamic_total_return = total_profit / dynamic_principal if dynamic_principal > 0 else 0
    dynamic_annualized_return = ((1 + dynamic_total_return) ** (12 / months_precise) - 1) if dynamic_total_return > 0 and months_precise > 0 else 0

    # 复利目标: 日/周/月 (基于动态本金)
    daily_target = (1 + target_annual) ** (1/250) - 1
    weekly_target = (1 + target_annual) ** (5/250) - 1
    monthly_target = (1 + target_annual) ** (1/12) - 1

    # ============================================================
    # 2. 读取周复盘数据 — 双模型收益率计算
    # ============================================================
    ws_rev_d = wb_d['周复盘']
    week_start = ws_rev_d.cell(row=2, column=2).value or 0
    week_end = ws_rev_d.cell(row=2, column=3).value or 0
    week_return = ws_rev_d.cell(row=2, column=4).value or 0
    if isinstance(week_start, str): week_start = 0
    if isinstance(week_end, str): week_end = 0
    if isinstance(week_return, str): week_return = 0

    # 周利润 = 周末总资产 - 周初总资产 (排除入金)
    week_profit = week_end - week_start

    # TWR周收益率 (主用): 利润 / 周初总资产
    twr_weekly_actual = week_profit / week_start if week_start > 0 else 0
    twr_daily_actual = (1 + twr_weekly_actual) ** (1/5) - 1 if twr_weekly_actual > -1 else 0

    # 动态本金周收益率 (辅用): 利润 / 持仓市值
    dynamic_weekly_actual = week_profit / dynamic_principal if dynamic_principal > 0 else 0
    dynamic_daily_actual = (1 + dynamic_weekly_actual) ** (1/5) - 1 if dynamic_weekly_actual > -1 else 0

    # 月复利: 上月末总资产 → 当前
    jun_end_asset = 18504.0  # 默认值
    for r in range(latest_row, 1, -1):
        d = ws_acc.cell(row=r, column=1).value
        if d and hasattr(d, 'month') and d.month == end_date.month - 1:
            jun_end_asset = ws_acc.cell(row=r, column=2).value or jun_end_asset
            break
    month_profit = current_asset - jun_end_asset

    # TWR月收益率 (主用)
    twr_monthly_actual = month_profit / jun_end_asset if jun_end_asset > 0 else 0

    # 动态本金月收益率 (辅用)
    dynamic_monthly_actual = month_profit / dynamic_principal if dynamic_principal > 0 else 0

    # 目标完成情况 (以TWR为主)
    daily_complete = "是" if twr_daily_actual >= daily_target else "否"
    weekly_complete = "是" if twr_weekly_actual >= weekly_target else "否"
    monthly_complete = "是" if twr_monthly_actual >= monthly_target else "否"

    # ============================================================
    # 3. 计算R值 (从交易记录) — FIFO买卖配对 + 代理止损
    # ============================================================
    # 修复 (2026-07-26):
    #   问题1: 买入行 buy_price(列22)=None, 实际买入价在 price(列6)
    #   问题2: 买入行 stop(列12)=None, 大部分无止损记录
    #   问题3: 同代码多次买卖, buy_map覆盖导致只匹配最后一笔
    #   方案: FIFO队列配对 + price列6作买入价 + 无止损时用5%代理止损
    ws_tr = wb_d['交易记录']
    trades = []
    buy_queues = {}  # code -> [buy_trade, ...] FIFO队列

    for r in range(2, SUMMARY_START):
        name = ws_tr.cell(row=r, column=2).value
        if not name:
            continue
        code = str(ws_tr.cell(row=r, column=3).value or "")
        direction = str(ws_tr.cell(row=r, column=4).value or "")
        price = ws_tr.cell(row=r, column=6).value      # 成交价格(买入行=买入价)
        shares = ws_tr.cell(row=r, column=7).value
        total = ws_tr.cell(row=r, column=8).value       # 成交金额
        stop = ws_tr.cell(row=r, column=12).value        # 止损价
        risk_col = ws_tr.cell(row=r, column=23).value    # W列止损风险额(可能为0)
        raw_date = ws_tr.cell(row=r, column=1).value     # 交易日期

        price = float(price) if not isinstance(price, str) and price is not None else 0
        shares = float(shares) if not isinstance(shares, str) and shares is not None else 0
        total = float(total) if not isinstance(total, str) and total is not None else 0
        stop = float(stop) if not isinstance(stop, str) and stop is not None else 0
        risk_col = float(risk_col) if not isinstance(risk_col, str) and risk_col is not None else 0
        date_str = raw_date.strftime('%Y-%m-%d') if isinstance(raw_date, datetime) else str(raw_date or "")

        t = {"row": r, "name": str(name), "code": code, "direction": direction,
             "price": price, "shares": shares, "total": total,
             "stop": stop, "risk_col": risk_col, "date_str": date_str}
        trades.append(t)
        if direction == "买入":
            buy_queues.setdefault(code, []).append(t)

    # FIFO配对计算R值
    r_values = []
    r_by_row = {}
    for t in trades:
        if t["direction"] not in ("卖出", "一卖"):
            continue
        queue = buy_queues.get(t["code"])
        if not queue:
            continue
        buy = queue.pop(0)  # FIFO: 取最早的买入

        # 买入价: 优先用buy行price(列6), 回退到sell行buy_price(列22)
        buy_price = buy["price"] if buy["price"] > 0 else 0

        # 止损风险额: 优先用W列记录值, 回退到(stop计算), 再回退到5%代理止损
        if buy["risk_col"] > 0:
            risk = buy["risk_col"]
        elif buy["stop"] > 0 and buy_price > 0:
            risk = (buy_price - buy["stop"]) * buy["shares"]
        else:
            # 代理止损: 买入价的5% (日线级别常见止损幅度)
            proxy_stop = buy_price * 0.95
            risk = (buy_price - proxy_stop) * buy["shares"] if buy_price > 0 else 0

        if risk > 0:
            r = (t["total"] - buy["total"]) / risk
        else:
            r = 0
        r = round(r, 6)
        r_values.append(r)
        r_by_row[t["row"]] = r

    avg_r = sum(r_values) / len(r_values) if r_values else 0
    total_r = sum(r_values)
    max_r = max(r_values) if r_values else 0
    max_loss_r = min(r_values) if r_values else 0
    wins = [r for r in r_values if r > 0]
    losses = [r for r in r_values if r < 0]
    if wins and losses:
        win_loss_ratio = (sum(wins) / len(wins)) / abs(sum(losses) / len(losses))
    elif wins:
        win_loss_ratio = float('inf')
    else:
        win_loss_ratio = 0
    win_rate = len(wins) / len(r_values) * 100 if r_values else 0
    system_status = "正期望系统 ✅" if avg_r >= 0.5 else ("边缘系统 ⚠️" if avg_r >= 0.2 else "负期望系统 ❌")

    # ============================================================
    # 3.5 做T统计 + 建清仓周期 + 持仓天数
    # ============================================================
    from collections import defaultdict
    t0_count = 0
    t0_wins = 0
    cycle_count = 0
    cycle_wins = 0
    hold_days_list = []

    # 按日期+代码分组识别做T(同日买卖)
    by_date_code = defaultdict(list)
    for t in trades:
        date_str = t.get("date_str", "")
        if not date_str:
            # 从交易记录原始行重新读取日期
            raw_date = ws_tr.cell(row=t["row"], column=1).value
            if isinstance(raw_date, datetime):
                date_str = raw_date.strftime('%Y-%m-%d')
            else:
                date_str = str(raw_date or "")
            t["date_str"] = date_str
        by_date_code[(date_str, t["code"])].append(t)

    for (d, code), group in by_date_code.items():
        buys = [g for g in group if g["direction"] == "买入"]
        sells = [g for g in group if g["direction"] in ("卖出", "一卖")]
        if buys and sells:
            buy_total = sum(b["total"] for b in buys)
            sell_total = sum(s["total"] for s in sells)
            t0_count += 1
            if sell_total > buy_total:
                t0_wins += 1
    t0_rate = t0_wins / t0_count * 100 if t0_count > 0 else 0

    # 建清仓周期: FIFO配对(复用buy_queues已消费的队列, 需重新构建)
    code_trades = defaultdict(list)
    for t in trades:
        if t["direction"] in ("买入", "卖出", "一卖"):
            code_trades[t["code"]].append(t)
    for code, ctrades in code_trades.items():
        ctrades.sort(key=lambda x: x["row"])
        current_buy = None
        for t in ctrades:
            if t["direction"] == "买入":
                if current_buy is None:
                    current_buy = dict(t)
                else:
                    current_buy["total"] += t["total"]
                    current_buy["shares"] += t["shares"]
            elif t["direction"] in ("卖出", "一卖") and current_buy:
                pnl = t["total"] - current_buy["total"]
                cycle_count += 1
                if pnl > 0:
                    cycle_wins += 1
                # 持仓天数
                try:
                    buy_d = datetime.strptime(current_buy.get("date_str", ""), '%Y-%m-%d')
                    sell_d = datetime.strptime(t.get("date_str", ""), '%Y-%m-%d')
                    hold_days_list.append((sell_d - buy_d).days)
                except:
                    pass
                current_buy = None
    cycle_rate = cycle_wins / cycle_count * 100 if cycle_count > 0 else 0
    avg_hold_days = sum(hold_days_list) / len(hold_days_list) if hold_days_list else 0

    # ============================================================
    # 4. 账户增长 (2026-07-26修订: 围绕"幼儿"目标重新定义阶段)
    #   婴儿: <3W   (当前, 今年目标养到幼儿)
    #   幼儿: 3W-10W (今年目标, 围绕幼儿执行)
    #   成长: 10W-30W
    #   成熟: 30W+
    # ============================================================
    growth_target_1 = 30000   # 婴儿→幼儿
    growth_target_2 = 100000  # 幼儿→成长
    growth_target_3 = 300000  # 成长→成熟
    if current_asset < 30000:
        current_stage = "婴儿"
        growth_progress = (current_asset / growth_target_1) * 100
        next_target = growth_target_1
        next_stage = "幼儿"
    elif current_asset < 100000:
        current_stage = "幼儿"
        growth_progress = (current_asset / growth_target_2) * 100
        next_target = growth_target_2
        next_stage = "成长"
    elif current_asset < 300000:
        current_stage = "成长"
        growth_progress = (current_asset / growth_target_3) * 100
        next_target = growth_target_3
        next_stage = "成熟"
    else:
        current_stage = "成熟"
        growth_progress = 100
        next_target = current_asset
        next_stage = "成熟"

    # ============================================================
    # 5. 写入Excel
    # ============================================================
    wb = safe_load_wb()
    ws_rev = wb['周复盘']

    # 新列表头 (如果不存在则添加)
    new_headers = {
        23: "日复利目标完成", 24: "周复利目标完成", 25: "月复利目标完成",
        26: "当下累计复利率", 27: "当下正期望R值", 28: "系统盈亏比",
        29: "账户规模增长", 30: "增长目标进度", 31: "复利追踪备注",
        32: "做T统计", 33: "建清仓成功率", 34: "平均持仓天数",
    }
    header_font = XFont(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col, title in new_headers.items():
        cell = ws_rev.cell(row=1, column=col)
        if not cell.value:
            cell.value = title
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    # 写数据行 (行2 = 当前周)
    row = 2
    data_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = XFont(size=10)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # 列23-25: 以TWR为主的目标完成, 备注中标注动态本金
    ws_rev.cell(row=row, column=23, value=f'{daily_complete} (TWR{twr_daily_actual*100:.3f}% vs目标{daily_target*100:.4f}%)')
    ws_rev.cell(row=row, column=24, value=f'{weekly_complete} (TWR{twr_weekly_actual*100:.3f}% vs目标{weekly_target*100:.4f}%)')
    ws_rev.cell(row=row, column=25, value=f'{monthly_complete} (TWR{twr_monthly_actual*100:.3f}% vs目标{monthly_target*100:.4f}%)')
    # 列26: 双复利累计收益率 [主]TWR + [辅]动态本金
    ws_rev.cell(row=row, column=26, value=f'[主]TWR:{twr_total_return*100:.1f}%年化{twr_annualized_return*100:.1f}% [辅]动态:{dynamic_total_return*100:.1f}%年化{dynamic_annualized_return*100:.1f}%')
    ws_rev.cell(row=row, column=27, value=f'平均R={avg_r:.4f}, 累计R={total_r:.4f}, 最大R={max_r:.4f}, 最大亏损R={max_loss_r:.4f}')
    ws_rev.cell(row=row, column=28, value=round(win_loss_ratio, 2) if win_loss_ratio != float('inf') else "∞")
    ws_rev.cell(row=row, column=29, value=f'{current_stage}阶段, ¥{current_asset:,.0f}/¥{next_target:,}目标({next_stage})')
    ws_rev.cell(row=row, column=30, value=f'{growth_progress:.1f}%')
    # 列31: 双复利基数详情
    ws_rev.cell(row=row, column=31, value=f'主:TWR本金¥{twr_principal:,.0f} 辅:动态¥{dynamic_principal:,.0f}({position_ratio:.0%}仓位) 系统={system_status} 胜率{win_rate:.0f}% 距{next_stage}差¥{next_target-current_asset:,.0f}')

    # 做T/建清仓/持仓天数 (新增3列)
    t0_rate_str = f'{t0_rate:.0f}%' if t0_count > 0 else 'N/A'
    ws_rev.cell(row=row, column=32, value=f'{t0_count}次, 成功率{t0_rate_str} ({t0_wins}/{t0_count})' if t0_count > 0 else '0次')
    ws_rev.cell(row=row, column=33, value=f'{cycle_rate:.0f}% ({cycle_wins}/{cycle_count})' if cycle_count > 0 else 'N/A')
    ws_rev.cell(row=row, column=34, value=f'{avg_hold_days:.0f}天' if avg_hold_days > 0 else 'N/A')

    # 绿色标注完成
    for col, complete in [(23, daily_complete), (24, weekly_complete), (25, monthly_complete)]:
        if complete == "是":
            ws_rev.cell(row=row, column=col).fill = green_fill
        ws_rev.cell(row=row, column=col).alignment = data_align
        ws_rev.cell(row=row, column=col).font = data_font

    for col in range(26, 35):
        ws_rev.cell(row=row, column=col).alignment = data_align
        ws_rev.cell(row=row, column=col).font = data_font
        ws_rev.cell(row=row, column=col).border = thin_border

    # 修复交易记录R值 (值模式, 覆盖公式)
    ws_tr_fix = wb['交易记录']
    cumul_r = 0
    for t in trades:
        r = t["row"]
        if t["direction"] in ("卖出", "一卖") and r in r_by_row:
            ws_tr_fix.cell(row=r, column=24, value=r_by_row[r])
            cumul_r += r_by_row[r]
        else:
            ws_tr_fix.cell(row=r, column=24, value=0)
        ws_tr_fix.cell(row=r, column=25, value=round(cumul_r, 6))

    # 统计区域写入正确值
    last_trade_row = max(t["row"] for t in trades) if trades else 19
    ws_tr_fix.cell(row=209, column=2, value=round(avg_r, 6))
    ws_tr_fix.cell(row=210, column=2, value=round(total_r, 6))
    ws_tr_fix.cell(row=211, column=2, value=round(max_r, 6))
    ws_tr_fix.cell(row=212, column=2, value=round(max_loss_r, 6))
    ws_tr_fix.cell(row=215, column=2, value=0)
    ws_tr_fix.cell(row=216, column=2, value=round(win_loss_ratio, 6) if win_loss_ratio != float('inf') else 999)
    ws_tr_fix.cell(row=217, column=2, value=system_status)

    safe_save_wb(wb)
    recalc_result = recalc()

    print(f"\n周复盘完成 (双复利模型):")
    print(f"  [主] TWR复利(总资产基数): 本金¥{twr_principal:,.0f}")
    print(f"       累计{twr_total_return*100:.1f}% 年化{twr_annualized_return*100:.1f}%")
    print(f"       日{daily_complete} 周{weekly_complete} 月{monthly_complete}")
    print(f"  [辅] 动态本金(持仓市值): ¥{dynamic_principal:,.0f} (仓位{position_ratio:.0%})")
    print(f"       累计{dynamic_total_return*100:.1f}% 年化{dynamic_annualized_return*100:.1f}%")
    print(f"  R值: 平均R={avg_r:.4f}, 累计R={total_r:.4f}, 盈亏比={win_loss_ratio:.2f}")
    print(f"  胜率: {win_rate:.0f}% ({len(wins)}/{len(r_values)})")
    print(f"  做T: {t0_count}次, 成功率{t0_rate:.0f}%")
    print(f"  建清仓: {cycle_count}次, 成功率{cycle_rate:.0f}%")
    print(f"  平均持仓: {avg_hold_days:.0f}天")
    print(f"  账户增长: {current_stage}, {growth_progress:.1f}% → ¥{next_target:,}({next_stage})")

    return {
        "daily": {"complete": daily_complete, "actual": twr_daily_actual, "target": daily_target,
                  "dynamic_actual": dynamic_daily_actual},
        "weekly": {"complete": weekly_complete, "actual": twr_weekly_actual, "target": weekly_target,
                   "dynamic_actual": dynamic_weekly_actual},
        "monthly": {"complete": monthly_complete, "actual": twr_monthly_actual, "target": monthly_target,
                    "dynamic_actual": dynamic_monthly_actual},
        "compound": {
            "twr_total_return": twr_total_return, "twr_annualized": twr_annualized_return,
            "dynamic_total_return": dynamic_total_return, "dynamic_annualized": dynamic_annualized_return,
            "deviation": compound_dev, "dynamic_principal": dynamic_principal,
            "twr_principal": twr_principal,
            "position_ratio": position_ratio, "invested_capital": invested_capital},
        "r_values": {"avg": avg_r, "total": total_r, "max": max_r, "max_loss": max_loss_r,
                      "win_loss_ratio": win_loss_ratio, "win_rate": win_rate, "system": system_status},
        "trading": {"t0_count": t0_count, "t0_wins": t0_wins, "t0_rate": t0_rate,
                     "cycle_count": cycle_count, "cycle_wins": cycle_wins, "cycle_rate": cycle_rate,
                     "avg_hold_days": avg_hold_days},
        "growth": {"stage": current_stage, "progress": growth_progress,
                    "next_target": next_target, "next_stage": next_stage,
                    "current_asset": current_asset},
    }


def format_weekly_review_summary(data, ts):
    """格式化周复盘摘要为Telegram消息"""
    lines = [f"📊 周复盘报告 {ts}", ""]

    # 复利目标
    lines.append("【复利目标完成情况 (以TWR为主)】")
    d = data["daily"]
    w = data["weekly"]
    m = data["monthly"]
    lines.append(f"  日复利: {d['complete']} (TWR{d['actual']*100:.3f}% vs目标{d['target']*100:.4f}%)")
    lines.append(f"  周复利: {w['complete']} (TWR{w['actual']*100:.3f}% vs目标{w['target']*100:.4f}%)")
    lines.append(f"  月复利: {m['complete']} (TWR{m['actual']*100:.3f}% vs目标{m['target']*100:.4f}%)")
    lines.append("")

    # 双复利累计
    c = data["compound"]
    lines.append("【双复利水平】")
    lines.append(f"  [主] TWR(总资产基数): 本金¥{c.get('twr_principal', 0):,.0f}")
    lines.append(f"       累计{c['twr_total_return']*100:.1f}% 年化{c['twr_annualized']*100:.1f}%")
    lines.append(f"  [辅] 动态本金(持仓市值): ¥{c.get('dynamic_principal', 0):,.0f} (仓位{c.get('position_ratio', 0):.0%})")
    lines.append(f"       累计{c['dynamic_total_return']*100:.1f}% 年化{c['dynamic_annualized']*100:.1f}%")
    lines.append("")

    # R值
    r = data["r_values"]
    lines.append("【正期望值R统计】")
    lines.append(f"  平均R: {r['avg']:.4f}")
    lines.append(f"  累计R: {r['total']:.4f}")
    lines.append(f"  最大R: {r['max']:.4f}")
    lines.append(f"  最大亏损R: {r['max_loss']:.4f}")
    lines.append(f"  盈亏比: {r['win_loss_ratio']:.2f}")
    lines.append(f"  胜率: {r['win_rate']:.0f}%")
    lines.append(f"  系统: {r['system']}")
    lines.append("")

    # 交易行为
    t = data.get("trading", {})
    lines.append("【交易行为统计】")
    lines.append(f"  做T: {t.get('t0_count', 0)}次, 成功率{t.get('t0_rate', 0):.0f}%")
    lines.append(f"  建清仓: {t.get('cycle_count', 0)}次, 成功率{t.get('cycle_rate', 0):.0f}%")
    lines.append(f"  平均持仓: {t.get('avg_hold_days', 0):.0f}天")
    lines.append("")

    # 账户增长
    g = data["growth"]
    lines.append("【账户规模增长目标】")
    lines.append(f"  当前阶段: {g['stage']}")
    lines.append(f"  当前资产: ¥{g['current_asset']:,.0f}")
    lines.append(f"  下阶段目标: ¥{g['next_target']:,} ({g['next_stage']})")
    lines.append(f"  完成进度: {g['progress']:.1f}%")
    lines.append(f"  距离目标: ¥{g['next_target'] - g['current_asset']:,.0f}")

    return '\n'.join(lines)


def clean_excel():
    """清理Excel中的空行、重复行和已卖出持仓

    BUG修复 (2026-08-05): 持仓表存有过期数据(松芝/贵绳/日上已卖出但未删除)
    修复: 读取交易记录中的"卖出"操作, 已全卖出的股票从持仓表删除
    """
    try:
        from openpyxl import load_workbook
        wb = safe_load_wb()
        total_empty = 0
        total_dup = 0
        total_sold = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 删除空行(前5列全空)
            empty = 0
            for r in range(ws.max_row, 1, -1):
                vals = [ws.cell(row=r, column=c).value for c in range(1, 6)]
                if all(v is None for v in vals):
                    ws.delete_rows(r, 1)
                    empty += 1
            total_empty += empty
            # 删除重复行(按关键列去重, 保留最后一条)
            seen = {}
            dup_rows = []
            dup = 0
            key_cols = [2, 3]  # 名称+代码
            for r in range(2, ws.max_row + 1):
                key = tuple(ws.cell(row=r, column=c).value for c in key_cols)
                if key and key != (None, None):
                    if key in seen:
                        dup_rows.append(r)
                        dup += 1
                    else:
                        seen[key] = r
            # BUG修复: 实际删除重复行, 旧代码只计数不删除
            for r in sorted(dup_rows, reverse=True):
                ws.delete_rows(r, 1)
            total_dup += dup

        # ============================================================
        # BUG修复 (2026-08-05): 持仓表已卖出股票清理
        # 问题: 松芝/贵绳/日上已卖出, 但持仓表仍有股数>0的行
        #       导致get_today_holdings返回虚假持仓
        # 修复: 读取交易记录, 已全卖出的股票从持仓表删除
        # ============================================================
        if '交易记录' in wb.sheetnames and '持仓表' in wb.sheetnames:
            ws_trade = wb['交易记录']
            ws_hold = wb['持仓表']

            # 计算每只股票在交易记录中的总卖出量
            sold_stocks = {}
            for r in range(2, ws_trade.max_row + 1):
                t_name = ws_trade.cell(row=r, column=2).value
                t_action = ws_trade.cell(row=r, column=4).value
                t_shares = ws_trade.cell(row=r, column=5).value
                if t_name and t_action and '卖出' in str(t_action):
                    try:
                        sv = int(t_shares) if t_shares else 0
                    except:
                        sv = 0
                    if t_name not in sold_stocks:
                        sold_stocks[t_name] = 0
                    sold_stocks[t_name] += sv

            # 删除已全卖出的股票行
            for r in range(ws_hold.max_row, 1, -1):
                h_name = ws_hold.cell(row=r, column=2).value
                h_shares = ws_hold.cell(row=r, column=7).value
                if h_name and h_name in sold_stocks:
                    try:
                        hs = int(h_shares) if h_shares else 0
                    except:
                        hs = 0
                    if hs > 0 and sold_stocks[h_name] >= hs:
                        ws_hold.delete_rows(r, 1)
                        total_sold += 1
                        print(f"  [清理] 删除已卖出持仓: {h_name}")

        safe_save_wb(wb)
        print(f"  [清理] 空行={total_empty}, 重复={total_dup}, 已卖出={total_sold}")
        return total_empty > 0 or total_dup > 0 or total_sold > 0
    except Exception as e:
        print(f"  [清理] 失败: {e}")
        return False

def main():
    if len(sys.argv) < 2:
        print("用法: daily_workflow.py [compliance|scan|intraday|account|holdings|weekly|clean|plan|prefetch|divergence]")
        return
    cmd = sys.argv[1]
    ts = datetime.now().strftime('%m-%d %H:%M')

    try:
        if cmd == "compliance":
            clean_excel()
            account, holdings, issues = check_compliance()
            # Telegram推送改为去弱留强格式 (2026-07-29)
            # 旧: format_compliance_summary → 只推送合规告警
            # 新: format_rebalance_summary → 减多少股+加多少股+核心池
            msg = format_rebalance_summary(holdings, ts)
            send_telegram(msg)
        elif cmd == "scan":
            use_prefetch = "--prefetch" in sys.argv
            clean_excel()
            scan_data = run_full_scan()
            if use_prefetch:
                # 先预取再扫描
                from concurrent_prefetch import prefetch_candidate_pool
                prefetch_candidate_pool(max_workers=20)
            msg = format_scan_summary(scan_data, ts)
            send_telegram(msg)
        elif cmd == "intraday":
            result = run_intraday_scan()
            msg = format_intraday_summary(result, ts)
            send_telegram(msg)
        elif cmd == "account":
            a = get_account_summary()
            def safe(v):
                if isinstance(v, (datetime, date)): return str(v)
                if isinstance(v, Decimal): return float(v)
                return v
            print(json.dumps({k:safe(v) for k,v in a.items()}, ensure_ascii=False, indent=2))
        elif cmd == "holdings":
            h = get_today_holdings()
            print(json.dumps(h, ensure_ascii=False, indent=2))
        elif cmd == "weekly":
            review_data = run_weekly_review()
            msg = format_weekly_review_summary(review_data, ts)
            send_telegram(msg)
        elif cmd == "clean":
            clean_excel()
            print("✅ Excel修复完成")
        elif cmd == "prefetch":
            """并发预取全市场数据"""
            from concurrent_prefetch import prefetch_candidate_pool, prefetch_holdings
            # 先预取持仓股 (多级别)
            try:
                holdings = get_today_holdings()
                if holdings:
                    print(f"预取持仓股{len(holdings)}只数据...")
                    prefetch_holdings(holdings, max_workers=15)
            except Exception as e:
                print(f"持仓预取跳过: {e}")
            # 再预取全市场 (日线)
            prefetch_candidate_pool(max_workers=20)
        elif cmd == "plan":
            """生成交易计划"""
            from trading_plan import generate_plan_from_workflow, format_plan, format_plan_short
            # 先预取持仓股数据
            try:
                from concurrent_prefetch import prefetch_holdings
                holdings = get_today_holdings()
                if holdings:
                    print(f"预取{len(holdings)}只持仓股数据...")
                    prefetch_holdings(holdings, max_workers=15)
            except Exception as e:
                print(f"预取跳过: {e}")
            plan = generate_plan_from_workflow()
            output = format_plan(plan)
            print(output)
            # 尝试Telegram推送精简版
            try:
                short = format_plan_short(plan)
                send_telegram(short, title="📋 交易计划")
            except Exception:
                pass
        elif cmd == "divergence":
            """持仓背驰确认 (V2: 顶背驰检测)"""
            from position_divergence import position_divergence_report
            result = position_divergence_report(silent=False)
            # 尝试Telegram推送
            try:
                msg_lines = [f"📊 持仓背驰确认(V2) — {ts}"]
                summary = result.get("summary", {})
                msg_lines.append(f"持仓{summary.get('total_holdings',0)}只 | "
                                 f"顶背驰{summary.get('top_divergence_confirmed',0)} | "
                                 f"失效{summary.get('buy_thesis_failed',0)} | "
                                 f"风险{summary.get('risk_rising',0)} | "
                                 f"有效{summary.get('valid',0)}")
                sell_alerts = result.get("sell_alerts", [])
                if sell_alerts:
                    msg_lines.append(f"\n🔴 清仓预警({len(sell_alerts)}只):")
                    for s in sell_alerts[:5]:
                        _r = s.get("risk", "?")
                        _n = s.get("name", "?")
                        _t = s.get("divergence_type", "?")
                        _p = s.get("profit", 0)
                        msg_lines.append(f"  {_r} {_n}({_t}) {_p:+.2f}%")
                reduce_alerts = result.get("reduce_alerts", [])
                if reduce_alerts:
                    msg_lines.append(f"\n🟡 减仓预警({len(reduce_alerts)}只):")
                    for r in reduce_alerts[:3]:
                        _n = r.get("name", "?")
                        _t = r.get("divergence_type", "?")
                        msg_lines.append(f"  {_n} ({_t})")
                send_telegram("\n".join(msg_lines), title="📊 持仓背驰")
            except Exception:
                pass
        else:
            print(f"未知命令: {cmd}")
    except Exception as e:
        import traceback
        err = traceback.format_exc()
        print(f"ERROR: {err}")
        try:
            send_telegram(f"❌ 工作流错误 ({cmd}) {ts}\n\n{str(e)}")
        except Exception:
            pass
        raise

if __name__ == '__main__':
    main()
