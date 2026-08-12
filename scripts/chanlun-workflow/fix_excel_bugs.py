#!/usr/bin/env python3
"""修复Excel数据问题: 空行+重复行+冗余数据"""
import sys, os
from openpyxl import load_workbook
from datetime import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
WB = os.path.join(SCRIPT_DIR, 'trade-workbook.xlsx')
BACKUP = os.path.join(SCRIPT_DIR, f'trade-workbook-backup-{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')

# 1. 备份
import shutil
shutil.copy2(WB, BACKUP)
print(f"[备份] 已创建: {os.path.basename(BACKUP)}")

wb = load_workbook(WB)

def remove_empty_rows(ws, key_col=1):
    """删除完全空行"""
    removed = 0
    for r in range(ws.max_row, 1, -1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column + 1, 10))]
        if all(v is None for v in row_vals):
            ws.delete_rows(r, 1)
            removed += 1
    return removed

def remove_duplicate_rows(ws, key_cols, keep='last'):
    """删除重复行, 保留最后一条"""
    seen = {}
    rows_to_delete = []
    for r in range(2, ws.max_row + 1):
        key = tuple(ws.cell(row=r, column=c).value for c in key_cols)
        if key and key != (None,) * len(key_cols):
            if key in seen:
                rows_to_delete.append(r)
            else:
                seen[key] = r
    rows_to_delete.sort(reverse=True)
    for r in rows_to_delete:
        ws.delete_rows(r, 1)
    return len(rows_to_delete)

# === 修复 持仓表 ===
ws = wb['持仓表']
print(f"\n[持仓表] 修复前行数: {ws.max_row}")

empty = remove_empty_rows(ws)
dup = remove_duplicate_rows(ws, key_cols=[2, 3, 7])  # 名称+代码+持股数
print(f"  删除空行: {empty}")
print(f"  删除重复: {dup}")
print(f"  修复后行数: {ws.max_row}")

# === 修复 交易记录 ===
ws = wb['交易记录']
print(f"\n[交易记录] 修复前行数: {ws.max_row}")

empty = remove_empty_rows(ws)
dup = remove_duplicate_rows(ws, key_cols=[1, 2, 3, 6, 7])  # 日期+名称+代码+价格+股数
print(f"  删除空行: {empty}")
print(f"  删除重复: {dup}")
print(f"  修复后行数: {ws.max_row}")

# === 修复 心态日志 ===
ws = wb['心态日志']
print(f"\n[心态日志] 修复前行数: {ws.max_row}")

empty = remove_empty_rows(ws)
dup = remove_duplicate_rows(ws, key_cols=[1, 2])  # 日期+操作类型
print(f"  删除空行: {empty}")
print(f"  删除重复: {dup}")
print(f"  修复后行数: {ws.max_row}")

# === 保存 ===
wb.save(WB)
print(f"\n✅ 修复完成! 已保存: {os.path.basename(WB)}")