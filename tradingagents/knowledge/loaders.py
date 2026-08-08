"""
知识库文档加载器 — 适配 .py / .md / .txt / .xlsx 等来源

每个加载器返回 KnowledgeDocument 列表，供 KnowledgeIngestionService 切分和入库。
"""
import os
import hashlib
import re
from dataclasses import dataclass, field
from typing import Optional


@dataclass
class KnowledgeDocument:
    """知识文档单元"""
    text: str
    source_type: str           # python_rule / excel / markdown / model_metadata / python_code
    source_path: str
    domain: str                # chanlun / trading_params / case_review / workflow / chanlun_ml
    metadata: dict = field(default_factory=dict)


class PythonFileLoader:
    """加载 Python 文件，按函数/类切分为知识单元"""

    # 提取函数和类的定义块
    FUNC_PATTERN = re.compile(
        r'^((?:def|class)\s+\w+.+?)(?=\n(?:def|class)\s+|\Z)',
        re.MULTILINE | re.DOTALL
    )
    # 提取文档字符串
    DOCSTRING_PATTERN = re.compile(r'"""(.+?)"""', re.DOTALL)

    def load(self, path: str, domain: str = "chanlun", rule_name: str = "") -> list[KnowledgeDocument]:
        if not os.path.isfile(path):
            return []

        with open(path, 'r', encoding='utf-8') as f:
            content = f.read()

        docs = []
        matches = self.FUNC_PATTERN.findall(content)
        for i, match in enumerate(matches):
            match = match.strip()
            if len(match) < 50:  # 跳过过短的片段
                continue

            # 提取文档字符串作为摘要
            doc_match = self.DOCSTRING_PATTERN.search(match)
            summary = doc_match.group(1).strip()[:200] if doc_match else ""

            text = f"{summary}\n\n{match}" if summary else match
            meta = {
                "source_path": path,
                "domain": domain,
                "rule_name": rule_name or self._extract_name(match),
            }
            docs.append(KnowledgeDocument(
                text=text,
                source_type="python_rule" if "rule" in domain else "python_code",
                source_path=path,
                domain=domain,
                metadata=meta,
            ))

        # 如果没有匹配到函数，把整个文件作为一个文档
        if not docs and len(content) > 100:
            docs.append(KnowledgeDocument(
                text=content[:5000],
                source_type="python_code",
                source_path=path,
                domain=domain,
                metadata={"source_path": path, "domain": domain},
            ))

        return docs

    def _extract_name(self, text: str) -> str:
        """提取函数或类名"""
        first_line = text.split('\n')[0]
        parts = first_line.split()
        if len(parts) >= 2:
            return parts[1].split('(')[0]
        return "unknown"


class MarkdownFileLoader:
    """加载 Markdown 文件，按标题切分"""

    HEADER_PATTERN = re.compile(r'^(#{1,6}\s+.+)$', re.MULTILINE)

    def load(self, path: str, domain: str = "workflow") -> list[KnowledgeDocument]:
        if not os.path.isfile(path):
            return []

        with open(path, 'r', encoding='utf-8') as f:
            content = f.read()

        docs = []
        sections = self.HEADER_PATTERN.split(content)
        current_header = ""
        for i, section in enumerate(sections):
            if self.HEADER_PATTERN.match(section):
                current_header = section.strip()
            elif len(section.strip()) > 50:
                text = f"{current_header}\n{section.strip()}" if current_header else section.strip()
                docs.append(KnowledgeDocument(
                    text=text,
                    source_type="markdown",
                    source_path=path,
                    domain=domain,
                    metadata={"source_path": path, "domain": domain, "section": current_header},
                ))

        if not docs and len(content) > 100:
            docs.append(KnowledgeDocument(
                text=content[:5000],
                source_type="markdown",
                source_path=path,
                domain=domain,
                metadata={"source_path": path, "domain": domain},
            ))

        return docs


class ExcelWorkbookLoader:
    """加载 Excel 工作簿，按 sheet + 语义行块切分"""

    def load(self, path: str, domain: str = "trading_params",
             target_sheets: list = None) -> list[KnowledgeDocument]:
        if not os.path.isfile(path):
            return []

        try:
            from openpyxl import load_workbook
        except ImportError:
            return []

        docs = []
        wb = load_workbook(path, data_only=True)
        for sheet_name in wb.sheetnames:
            if target_sheets and sheet_name not in target_sheets:
                continue

            ws = wb[sheet_name]
            if ws.max_row < 2:
                continue

            # 读取表头
            headers = []
            for col in range(1, min(ws.max_column + 1, 20)):
                val = ws.cell(row=1, column=col).value
                headers.append(str(val) if val else f"col_{col}")

            # 按行块切分（每5行一个chunk）
            chunk_size = 5
            for start_row in range(2, ws.max_row + 1, chunk_size):
                end_row = min(start_row + chunk_size - 1, ws.max_row)
                lines = []
                for r in range(start_row, end_row + 1):
                    row_data = []
                    for col in range(1, len(headers) + 1):
                        val = ws.cell(row=r, column=col).value
                        row_data.append(str(val) if val is not None else "")
                    if any(row_data):
                        lines.append(" | ".join(f"{h}: {v}" for h, v in zip(headers, row_data)))

                if not lines:
                    continue

                text = f"[Sheet: {sheet_name}]\n" + "\n".join(lines)
                meta = {
                    "source_path": path,
                    "sheet_name": sheet_name,
                    "row_start": start_row,
                    "row_end": end_row,
                    "domain": domain,
                }
                docs.append(KnowledgeDocument(
                    text=text,
                    source_type="excel",
                    source_path=path,
                    domain=domain,
                    metadata=meta,
                ))

        wb.close()
        return docs


class ModelMetadataLoader:
    """加载模型 pkl 的元数据（不 embedding 二进制内容）"""

    def load(self, path: str, model_name: str = "", model_type: str = "MLP",
             domain: str = "chanlun_ml") -> list[KnowledgeDocument]:
        if not os.path.isfile(path):
            return []

        # 计算文件 hash
        content_hash = hashlib.sha256()
        with open(path, 'rb') as f:
            while chunk := f.read(8192):
                content_hash.update(chunk)

        # 尝试加载模型获取结构信息
        model_info = {}
        try:
            import pickle
            with open(path, 'rb') as f:
                obj = pickle.load(f)
            if hasattr(obj, 'get_params'):
                model_info = {k: str(v) for k, v in obj.get_params().items()}
            elif hasattr(obj, 'n_features_in_'):
                model_info['n_features'] = obj.n_features_in_
            if hasattr(obj, 'classes_'):
                model_info['classes'] = str(obj.classes_)
        except Exception:
            pass

        text = (
            f"Model: {model_name}\n"
            f"Type: {model_type}\n"
            f"File: {os.path.basename(path)}\n"
            f"Size: {os.path.getsize(path)} bytes\n"
            f"SHA256: {content_hash.hexdigest()[:16]}\n"
        )
        for k, v in model_info.items():
            text += f"{k}: {v}\n"

        meta = {
            "source_path": path,
            "model_name": model_name,
            "model_type": model_type,
            "domain": domain,
            "content_hash": content_hash.hexdigest()[:16],
        }
        return [KnowledgeDocument(
            text=text,
            source_type="model_metadata",
            source_path=path,
            domain=domain,
            metadata=meta,
        )]


def load_all_sources(source_root: str) -> dict[str, list[KnowledgeDocument]]:
    """
    从源目录加载所有知识文档，返回 {collection_name: [documents]} 映射
    """
    py_loader = PythonFileLoader()
    md_loader = MarkdownFileLoader()
    xlsx_loader = ExcelWorkbookLoader()
    model_loader = ModelMetadataLoader()

    result = {
        "kb_chanlun_rules": [],
        "kb_trade_playbook": [],
        "kb_case_review": [],
        "kb_workflow_docs": [],
    }

    # 1. 缠论规则 — beichi_analyzer.py
    beichi_path = os.path.join(source_root, "beichi_analyzer.py")
    if os.path.isfile(beichi_path):
        docs = py_loader.load(beichi_path, domain="chanlun", rule_name="beichi_detection")
        result["kb_chanlun_rules"].extend(docs)

    # 2. 交易手册 — trade-workbook.xlsx 关键变量表 + 执行清单
    xlsx_path = os.path.join(source_root, "trade-workbook.xlsx")
    if os.path.isfile(xlsx_path):
        docs = xlsx_loader.load(
            xlsx_path, domain="trading_params",
            target_sheets=["关键变量表", "执行清单"]
        )
        result["kb_trade_playbook"].extend(docs)

        # 3. 历史案例 — 交易记录 + 周复盘 + 心态日志
        case_docs = xlsx_loader.load(
            xlsx_path, domain="case_review",
            target_sheets=["交易记录", "周复盘", "心态日志", "候选池历史"]
        )
        result["kb_case_review"].extend(case_docs)

    # 4. 工作流文档 — daily_workflow.py + README
    workflow_path = os.path.join(source_root, "daily_workflow.py")
    if os.path.isfile(workflow_path):
        docs = py_loader.load(workflow_path, domain="workflow")
        result["kb_workflow_docs"].extend(docs)

    readme_path = os.path.join(source_root, "README.md")
    if os.path.isfile(readme_path):
        docs = md_loader.load(readme_path, domain="workflow")
        result["kb_workflow_docs"].extend(docs)

    # 5. 模型元数据
    for pkl_name, model_name in [("dl_model.pkl", "dl_model"), ("ep_model.pkl", "ep_model")]:
        pkl_path = os.path.join(source_root, pkl_name)
        if os.path.isfile(pkl_path):
            docs = model_loader.load(pkl_path, model_name=model_name, model_type="MLP")
            result["kb_chanlun_rules"].extend(docs)

    # 6. WORKFLOW.md → kb_workflow_docs
    workflow_md_path = os.path.join(source_root, "WORKFLOW.md")
    if os.path.isfile(workflow_md_path):
        docs = md_loader.load(workflow_md_path, domain="workflow")
        result["kb_workflow_docs"].extend(docs)

    # 7. Markdown 计划/复盘文件 → kb_case_review
    import glob
    case_md_patterns = ["周复盘_*.md", "去弱留强计划_*.md", "今日计划_*.md", "明日计划_*.md", "周*计划_*.md"]
    for pattern in case_md_patterns:
        for md_path in glob.glob(os.path.join(source_root, pattern)):
            docs = md_loader.load(md_path, domain="case_review")
            result["kb_case_review"].extend(docs)

    # 8. 持仓换股策略文档 → kb_trade_playbook
    for md_path in glob.glob(os.path.join(source_root, "持仓换股策略_*.md")):
        docs = md_loader.load(md_path, domain="trading_params")
        result["kb_trade_playbook"].extend(docs)

    return result
